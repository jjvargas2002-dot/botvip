from functools import wraps
from io import BytesIO
import csv
import io
import requests
from datetime import datetime, timedelta

def obtener_hora_peru():
    return datetime.utcnow() - timedelta(hours=5)

from flask import Flask, render_template, request, send_file, redirect, url_for, session, flash, jsonify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from werkzeug.security import generate_password_hash, check_password_hash
from config import Config
from models import db, Cliente, Caso, Operador, Averia, StockBranch, Box

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

MATERIALES_MASTER = [
    # Sección OLT
    {"codigo": "291368", "nombre": "OLT GPON ZTE C610 - 16 puertos DC", "seccion": "OLT"},
    {"codigo": "294701", "nombre": "Transceptor óptico bidireccional monofibra SFP GPON-OLT clase C+", "seccion": "OLT"},
    {"codigo": "283866", "nombre": "Módulo óptico bidireccional de doble fibra y canal único ZTE 10Km/1.25Gb", "seccion": "OLT"},
    # Sección ODF
    {"codigo": "294702", "nombre": "Patch Cord SC/UPC-SC/UPC 3m", "seccion": "ODF"},
    {"codigo": "89", "nombre": "Patch Cord LC/UPC-LC/UPC 3m", "seccion": "ODF"},
    {"codigo": "87", "nombre": "Patch Cord LC/UPC-LC/UPC 15m", "seccion": "ODF"},
    {"codigo": "86", "nombre": "Patch Cord LC/UPC-LC/UPC 10m", "seccion": "ODF"},
    {"codigo": "267990", "nombre": "ODF 48 puertos SC/UPC", "seccion": "ODF"},
    {"codigo": "2925", "nombre": "ODF 24 puertos SC/UPC", "seccion": "ODF"},
    {"codigo": "265191", "nombre": "Gabinete exterior para OLT y ODF", "seccion": "ODF"},
    # Sección CAJAS
    {"codigo": "299798", "nombre": "Caja de empalme tipo domo 48Fo, 8 puertos", "seccion": "CAJAS"},
    {"codigo": "299799", "nombre": "Caja de empalme tipo domo 24Fo, 8 puertos", "seccion": "CAJAS"},
    {"codigo": "424", "nombre": "Caja de empalme plana 24Fo, 4 puertos", "seccion": "CAJAS"},
    {"codigo": "423", "nombre": "Caja de empalme plana 12Fo, 4 puertos", "seccion": "CAJAS"},
    {"codigo": "8810", "nombre": "Caja de empalme plana 4Fo, 3 puertos", "seccion": "CAJAS"},
    {"codigo": "262079", "nombre": "Caja de empalme plana 1Fo, 2 puertos, empalme mecánico", "seccion": "CAJAS"},
    {"codigo": "300476", "nombre": "ODB, 11 puertos para adaptadores (HUB-BOX 70/30)", "seccion": "CAJAS"},
    {"codigo": "300477", "nombre": "ODB, 11 puertos para adaptadores (SUB-BOX 70/30)", "seccion": "CAJAS"},
    {"codigo": "300478", "nombre": "ODB, 11 puertos para adaptadores (SUB-BOX 50/50)", "seccion": "CAJAS"},
    {"codigo": "300479", "nombre": "ODB, 10 puertos para adaptadores (END-BOX 50/50)", "seccion": "CAJAS"},
    {"codigo": "298476", "nombre": "ODB, 9 puertos para adaptadores (EXP-BOX 50/50)", "seccion": "CAJAS"},
    {"codigo": "298912", "nombre": "ODB, 9 puertos para adaptadores (Edificio)", "seccion": "CAJAS"},
    {"codigo": "296344", "nombre": "ZTE_ODB, 2 puertos de fibra para adaptadores (Caja de unión)", "seccion": "CAJAS"},
    {"codigo": "294765", "nombre": "ZTE_ODB, 8 puertos MPO y 4 puertos de fibra para adaptadores (X-BOX)", "seccion": "CAJAS"},
    {"codigo": "294764", "nombre": "ZTE_ODB, 5 puertos para adaptadores (HUB-BOX)", "seccion": "CAJAS"},
    {"codigo": "294767", "nombre": "ZTE_ODB, 11 puertos para adaptadores (SUB-BOX 70/30)", "seccion": "CAJAS"},
    {"codigo": "294768", "nombre": "ZTE_ODB, 11 puertos para adaptadores (SUB-BOX 50/50)", "seccion": "CAJAS"},
    {"codigo": "294766", "nombre": "ZTE_ODB, 10 puertos para adaptadores (END-BOX 50/50)", "seccion": "CAJAS"},
    {"codigo": "296345", "nombre": "ZTE_ODB, 9 puertos para adaptadores (EXP-BOX 50/50)", "seccion": "CAJAS"},
    # Sección de Conectores
    {"codigo": "299379", "nombre": "Conector Waterproof", "seccion": "CONECTORES"},
    {"codigo": "299378", "nombre": "Conector rápido SC/APC", "seccion": "CONECTORES"},
    {"codigo": "Sin Código", "nombre": "Splitter PLC 1x8 (Pigtail PLC 1x8)", "seccion": "CONECTORES"},
    # Sección de Cables
    {"codigo": "43978", "nombre": "Fibra óptica ADSS 48 fibras, SPAN 100m", "seccion": "CABLES"},
    {"codigo": "63004", "nombre": "Fibra óptica ADSS 24 fibras, SPAN 200m", "seccion": "CABLES"},
    {"codigo": "67", "nombre": "Fibra óptica ADSS 24 fibras, SPAN 100m", "seccion": "CABLES"},
    {"codigo": "299344", "nombre": "Fibra óptica ASU 12Fo, SPAN 100m", "seccion": "CABLES"},
    {"codigo": "300378", "nombre": "Fibra óptica ASU 4Fo, SPAN 100m", "seccion": "CABLES"},
    {"codigo": "9266", "nombre": "Fibra óptica ASU 4Fo, SPAN 100m - Flexible", "seccion": "CABLES"},
    {"codigo": "300379", "nombre": "Fibra óptica ASU 1Fo, SPAN 100m", "seccion": "CABLES"},
    {"codigo": "294795", "nombre": "Fibra óptica preconectorizada de 12 núcleos 100m MPO/APC", "seccion": "CABLES"},
    {"codigo": "294796", "nombre": "Fibra óptica preconectorizada de 12 núcleos 200m MPO/APC", "seccion": "CABLES"},
    {"codigo": "296347", "nombre": "Fibra óptica preconectorizada de 12 núcleos 300m MPO/APC", "seccion": "CABLES"},
    {"codigo": "294797", "nombre": "Fibra óptica preconectorizada de 12 núcleos 500m MPO/APC", "seccion": "CABLES"},
    {"codigo": "294798", "nombre": "Fibra óptica preconectorizada de 12 núcleos 600m MPO/APC", "seccion": "CABLES"},
    {"codigo": "294799", "nombre": "Fibra óptica preconectorizada de 12 núcleos 700m MPO/APC", "seccion": "CABLES"},
    {"codigo": "294790", "nombre": "Fibra óptica preconectorizada de 1 núcleo 50m SC/APC", "seccion": "CABLES"},
    {"codigo": "294791", "nombre": "Fibra óptica preconectorizada de 1 núcleo 100m SC/APC", "seccion": "CABLES"},
    {"codigo": "294792", "nombre": "Fibra óptica preconectorizada de 1 núcleo 150m SC/APC", "seccion": "CABLES"},
    {"codigo": "294793", "nombre": "Fibra óptica preconectorizada de 1 núcleo 200m SC/APC", "seccion": "CABLES"},
    {"codigo": "294794", "nombre": "Fibra óptica preconectorizada de 1 núcleo 250m SC/APC", "seccion": "CABLES"},
    {"codigo": "295995", "nombre": "Fibra óptica preconectorizada de 1 núcleo 300m SC/APC", "seccion": "CABLES"},
    {"codigo": "299381", "nombre": "Fibra óptica 1 hilo - Drop", "seccion": "CABLES"},
    # Sección de Accesorios
    {"codigo": "296348", "nombre": "ZTE_Patch Cord SC/APC 7m", "seccion": "ACCESORIOS"},
    {"codigo": "306", "nombre": "Grapa de tensión para SPAN de 200m", "seccion": "ACCESORIOS"},
    {"codigo": "305", "nombre": "Grapa de tensión para SPAN de 100m", "seccion": "ACCESORIOS"},
    {"codigo": "313", "nombre": "Grapa de suspensión para SPAN de 100m", "seccion": "ACCESORIOS"},
    {"codigo": "350", "nombre": "Abrazaderas para cable OPGW", "seccion": "ACCESORIOS"},
    {"codigo": "298983", "nombre": "Retención preformada plástica para cable de 5mm-8mm", "seccion": "ACCESORIOS"},
    {"codigo": "294929", "nombre": "Retención preformada para cable de 6.8mm (Cable MPO)", "seccion": "ACCESORIOS"},
    {"codigo": "294928", "nombre": "Retención preformada para cable de 5mm (Cable preconectorizado)", "seccion": "ACCESORIOS"},
    {"codigo": "299380", "nombre": "Templador para cable Drop", "seccion": "ACCESORIOS"},
    {"codigo": "28497", "nombre": "Clevis tipo D", "seccion": "ACCESORIOS"},
    {"codigo": "294898", "nombre": "Clevis tipo trébol", "seccion": "ACCESORIOS"},
    {"codigo": "285043", "nombre": "Brazo de soporte 1.0m", "seccion": "ACCESORIOS"},
    {"codigo": "295136", "nombre": "Brazo de soporte 0.6m", "seccion": "ACCESORIOS"},
    {"codigo": "24941", "nombre": "Cruceta 60cm", "seccion": "ACCESORIOS"},
    {"codigo": "295280", "nombre": "Cable de acero 4mm, con recubrimiento PVC", "seccion": "ACCESORIOS"},
    {"codigo": "295360", "nombre": "Abrazadera colgante dieléctrica", "seccion": "ACCESORIOS"},
    {"codigo": "20910", "nombre": "Candado de acero 3/8\" para cable de acero", "seccion": "ACCESORIOS"},
    {"codigo": "8332", "nombre": "Cinta Bandit + juego de hebillas", "seccion": "ACCESORIOS"},
    {"codigo": "294899", "nombre": "Fleje de acero para postes con diámetros de 100-200mm", "seccion": "ACCESORIOS"},
    {"codigo": "295660", "nombre": "Cinta Bandit", "seccion": "ACCESORIOS"},
    {"codigo": "295661", "nombre": "Juego de hebillas", "seccion": "ACCESORIOS"},
    {"codigo": "Sin Código", "nombre": "Etiqueta para fibra óptica", "seccion": "ACCESORIOS"},
    {"codigo": "9610", "nombre": "Tubo corrugado de PVC ignífugo 25mm", "seccion": "ACCESORIOS"},
    {"codigo": "Sin Código", "nombre": "Abrazadera tipo oreja para cable", "seccion": "ACCESORIOS"},
    {"codigo": "Sin Código", "nombre": "Cintillo para cable 4x200mm", "seccion": "ACCESORIOS"}
]

def obtener_material_mapeado(codigo, nombre):
    nombre_lower = nombre.lower()
    
    # 1. Buscar en catálogo completo para obtener la sección si es necesario
    master_item = next((m for m in MATERIALES_MASTER if m["codigo"] == codigo or m["nombre"].lower() == nombre_lower), None)
    seccion = master_item["seccion"] if master_item else ""
    
    # 2. Cable Drop (código 283866)
    if (codigo in ["283866", "299381"] or "drop" in nombre_lower) and not ("templador" in nombre_lower or "grapa" in nombre_lower or "anclaje" in nombre_lower or "retencion" in nombre_lower or "retención" in nombre_lower or "soporte" in nombre_lower or "preformada" in nombre_lower or "tencion" in nombre_lower or "tensión" in nombre_lower or "abrazadera" in nombre_lower):
        return "283866", "Cable Drop", "CABLES"
        
    # 3. FAC (código 299378)
    elif codigo == "299378" or "fac" in nombre_lower or "conector rápido" in nombre_lower or "conector rapido" in nombre_lower:
        return "299378", "FAC", "CONECTORES"
        
    # 4. Waterproof (código 299379)
    elif codigo == "299379" or "waterproof" in nombre_lower:
        return "299379", "Waterproof", "CONECTORES"
        
    # 5. Mufas (código 299799): agrupar todos los códigos de la sección "CAJAS"
    elif seccion == "CAJAS" or "empalme" in nombre_lower or "caja de empalme" in nombre_lower or codigo in ["299798", "299799", "424", "423", "8810", "262079"] or "odb" in nombre_lower:
        return "299799", "Mufas", "CAJAS"
        
    # 6. Preconectorizado (código 294790): agrupar todas las fibras que digan "Preconectorizado" de la "Sección CABLES" (excluyendo preformadas/retenciones/grapas)
    elif ((seccion == "CABLES" and "preconectoriza" in nombre_lower) or "preconectoriza" in nombre_lower) and not ("preformada" in nombre_lower or "retencion" in nombre_lower or "retención" in nombre_lower or "grapa" in nombre_lower):
        return "294790", "Preconectorizado", "CABLES"
        
    # Fallback to original
    return codigo, nombre, seccion

@app.context_processor
def utility_processor():
    materiales_por_seccion = {}
    
    # Materiales comunes
    materiales_comunes_nombres = [
        "Fibra óptica preconectorizada de 12 núcleos 200m MPO/APC",
        "Caja de empalme plana 12Fo, 4 puertos",
        "Tubo corrugado de PVC ignífugo 25mm",
        "Conector Waterproof",
        "Grapa de tensión para SPAN de 200m",
        "Grapa de tensión para SPAN de 100m"
    ]
    
    comunes = []
    for nombre_comun in materiales_comunes_nombres:
        found = next((m for m in MATERIALES_MASTER if m["nombre"] == nombre_comun), None)
        if found:
            comunes.append(found)
            
    if comunes:
        materiales_por_seccion["MATERIALES COMUNES"] = comunes
        
    for mat in MATERIALES_MASTER:
        sec = mat["seccion"]
        if sec not in materiales_por_seccion:
            materiales_por_seccion[sec] = []
        materiales_por_seccion[sec].append(mat)
        
    return dict(materiales_por_seccion=materiales_por_seccion, MATERIALES_MASTER=MATERIALES_MASTER)

def sincronizar_sites():
    url_sites = "https://docs.google.com/spreadsheets/d/1eaNxCpm8JF1JcZS3_ldwMRINGYFaW6RsQQWybvRi_P8/export?format=csv&gid=894046404"
    try:
        response = requests.get(url_sites, timeout=15)
        if response.status_code != 200:
            return False, f"Error de conexión con pestaña SITES (Status: {response.status_code})"
        
        content = response.content.decode('utf-8-sig')
        f = io.StringIO(content)
        reader = csv.reader(f)
        
        header = None
        indices = {}
        for row in reader:
            row_upper = [col.strip().upper() for col in row]
            if "BRANCH" in row_upper and "SITE LOGICAL" in row_upper:
                header = row_upper
                indices = {col.strip().upper(): i for i, col in enumerate(row)}
                break
                
        if not header:
            return False, "Faltan columnas Branch o Site Logical en la pestaña SITES"
            
        branch_idx = indices.get("BRANCH")
        site_idx = indices.get("SITE LOGICAL")
            
        sites_list = []
        for row in reader:
            if not row or len(row) <= max(branch_idx, site_idx):
                continue
            branch = row[branch_idx].strip().upper()
            site = row[site_idx].strip().upper()
            if branch and site and site != "SITE LOGICAL":
                sites_list.append({
                    "branch": branch,
                    "site": site
                })
        
        import os
        import json
        static_dir = os.path.join(app.root_path, "static")
        os.makedirs(static_dir, exist_ok=True)
        with open(os.path.join(static_dir, "sites.json"), "w", encoding="utf-8") as out:
            json.dump(sites_list, out, ensure_ascii=False, indent=2)
            
        return True, f"Sincronizados {len(sites_list)} sites."
    except Exception as e:
        print("Error en sincronización de sites:", e)
        return False, f"Error en sites: {str(e)}"


def sincronizar_boxes():
    url_boxes = "https://docs.google.com/spreadsheets/d/1eaNxCpm8JF1JcZS3_ldwMRINGYFaW6RsQQWybvRi_P8/export?format=csv&gid=1980704985"
    try:
        response = requests.get(url_boxes, timeout=30)
        if response.status_code != 200:
            return False, f"Error de conexión con pestaña List of Boxes (Status: {response.status_code})"
        
        content = response.content.decode('utf-8-sig')
        f = io.StringIO(content)
        reader = csv.reader(f)
        
        # Skip report headers to find actual headers
        header = None
        indices = {}
        for row in reader:
            row_upper = [col.strip().upper() for col in row]
            if "NODE CODE" in row_upper or "LATITUDE" in row_upper:
                header = row
                indices = {col.strip().upper(): i for i, col in enumerate(row)}
                break
                
        if not header:
            return False, "Faltan columnas críticas en la pestaña List of Boxes"
            
        node_idx = indices.get("NODE CODE")
        lat_idx = indices.get("LATITUDE")
        lng_idx = indices.get("LONGITUDE")
        
        if node_idx is None or lat_idx is None or lng_idx is None:
            return False, "Columnas Node code, Latitude o Longitude no encontradas en List of Boxes"
            
        box_mappings = []
        for row in reader:
            if not row or len(row) <= max(node_idx, lat_idx, lng_idx):
                continue
            caja_val = row[node_idx].strip().upper()
            lat_val = row[lat_idx].strip().replace(",", ".").strip()
            lng_val = row[lng_idx].strip().replace(",", ".").strip()
            
            if caja_val and lat_val and lng_val:
                box_mappings.append({
                    "caja": caja_val,
                    "latitud": lat_val,
                    "longitud": lng_val
                })
        
        if box_mappings:
            db.session.query(Box).delete()
            db.session.bulk_insert_mappings(Box, box_mappings)
            db.session.commit()
            
        return True, f"Sincronizados {len(box_mappings)} boxes."
    except Exception as e:
        db.session.rollback()
        print("Error en sincronización de boxes:", e)
        return False, f"Error en boxes: {str(e)}"


def login_requerido(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "operador_id" not in session:
            flash("Por favor, inicia sesión para acceder al sistema.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def obtener_estadisticas(branch=None, es_admin=False):
    try:
        query_total = Averia.query
        query_pendientes = Averia.query.filter_by(estado="PENDIENTE")
        query_reparados = Averia.query.filter_by(estado="REPARADO")

        if not es_admin and branch and branch != "ALL":
            query_total = query_total.filter_by(branch=branch)
            query_pendientes = query_pendientes.filter_by(branch=branch)
            query_reparados = query_reparados.filter_by(branch=branch)

        return {
            "totales": query_total.count(),
            "pendientes": query_pendientes.count(),
            "reparados": query_reparados.count(),
        }
    except Exception as e:
        print("Error obteniendo estadísticas:", e)
        return {
            "totales": 0,
            "pendientes": 0,
            "reparados": 0,
        }



def asegurar_esquema():
    # Asegurar que las columnas existan dinámicamente si las tablas ya existen en el DB
    columnas = [
        "ALTER TABLE operadores ADD COLUMN IF NOT EXISTS dni VARCHAR(20) UNIQUE",
        "ALTER TABLE operadores ADD COLUMN IF NOT EXISTS rol VARCHAR(20) DEFAULT 'operador'",
        "ALTER TABLE operadores ADD COLUMN IF NOT EXISTS branch VARCHAR(30) DEFAULT 'ALL'",
        "ALTER TABLE operadores ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE",
        "ALTER TABLE operadores ADD COLUMN IF NOT EXISTS fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        
        # Columnas de averias
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS branch VARCHAR(50)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS codigo_wo VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS cuenta VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS detalles TEXT",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS dias_pendientes DOUBLE PRECISION",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS estado VARCHAR(50) DEFAULT 'PENDIENTE'",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS status_caja VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS contrata VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS periodo_pendiente VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS site VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS caja VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS coordenadas VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS origen VARCHAR(20) DEFAULT 'SHEETS'",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS fecha_resolucion TIMESTAMP",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS tecnico_id INTEGER REFERENCES operadores(id)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS material_cable_m INTEGER DEFAULT 0",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS material_conectores INTEGER DEFAULT 0",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS material_rosetas INTEGER DEFAULT 0",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS material_mangas INTEGER DEFAULT 0",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS material_acopladores INTEGER DEFAULT 0",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS material_comentarios TEXT",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS materiales_json TEXT",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS tipificacion VARCHAR(100)",
        "ALTER TABLE averias ADD COLUMN IF NOT EXISTS cuentas_asociadas TEXT",
        """
        CREATE TABLE IF NOT EXISTS stock_branch (
            id SERIAL PRIMARY KEY,
            branch VARCHAR(50) NOT NULL,
            material_codigo VARCHAR(50) NOT NULL,
            material_nombre VARCHAR(255) NOT NULL,
            stock_actual INTEGER DEFAULT 0,
            stock_enviado_noc INTEGER DEFAULT 0,
            fecha_envio_noc DATE,
            fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(branch, material_nombre)
        )
        """,
        "ALTER TABLE stock_branch DROP CONSTRAINT IF EXISTS uq_branch_material",
        "ALTER TABLE stock_branch DROP CONSTRAINT IF EXISTS stock_branch_branch_material_codigo_key",
        "ALTER TABLE stock_branch ADD CONSTRAINT uq_branch_material_nombre UNIQUE (branch, material_nombre)",
        "ALTER TABLE averias DROP CONSTRAINT IF EXISTS averias_cuenta_key",
        "ALTER TABLE averias ALTER COLUMN cuenta DROP NOT NULL"
    ]

    for consulta in columnas:
        try:
            db.session.execute(text(consulta))
        except Exception as e:
            print(f"Error ejecutando consulta de esquema {consulta}: {e}")
            db.session.rollback()
    db.session.commit()

    # Cleanup: Delete only non-ODN tickets (e.g. status ONLINE, contact problem) that have no technician or materials
    try:
        def es_ticket_no_odn_local(detalles, status_ont, status_caja, material_comentarios=None):
            det = (detalles or "").lower()
            ont = (status_ont or "").lower()
            caja = (status_caja or "").lower()
            com = (material_comentarios or "").lower()
            
            def contains_word(text, word):
                if word not in text:
                    return False
                if word == "online":
                    if "not online" in text or "no online" in text or "not_online" in text:
                        import re
                        occurrences = [m.start() for m in re.finditer("online", text)]
                        for start in occurrences:
                            pre_text = text[max(0, start-5):start]
                            if "not" not in pre_text and "no" not in pre_text:
                                return True
                        return False
                return True

            if contains_word(det, "online") or contains_word(ont, "online") or contains_word(caja, "online") or contains_word(com, "estado: online"):
                return True
            if "contacto" in det or "contacto" in ont or "contacto" in caja or "estado: contacto" in com:
                return True
            if "desea" in det or "desea" in ont or "desea" in caja or "desea" in com:
                return True
            if "contesta" in det or "contesta" in ont or "contesta" in caja or "contesta" in com:
                return True
            if "baja" in det or "baja" in ont or "baja" in caja or "baja" in com:
                return True
            return False

        all_avs = Averia.query.all()
        deleted_count = 0
        for av in all_avs:
            status_for_check = av.estado if av.estado == "ONLINE" else ""
            if es_ticket_no_odn_local(av.detalles, status_for_check, av.status_caja, av.material_comentarios):
                db.session.delete(av)
                deleted_count += 1
        db.session.commit()
        if deleted_count > 0:
            print(f"Cleaned up {deleted_count} non-ODN tickets from database.")
    except Exception as e:
        print("Error executing database cleanup:", e)
        db.session.rollback()


def crear_operador_defecto():
    try:
        # 1. Crear/restablecer Administrador general FBB
        admin = Operador.query.filter(db.func.lower(Operador.dni) == "fbb").first()
        if admin:
            admin.dni = "FBB"
            admin.password_hash = generate_password_hash("Bitel@123")
            admin.rol = "admin"
            admin.branch = "ALL"
            admin.correo = "admin@botvip.com"
        else:
            default_pwd = generate_password_hash("Bitel@123")
            admin = Operador(
                nombre="Administrador FBB",
                dni="FBB",
                correo="admin@botvip.com",
                password_hash=default_pwd,
                rol="admin",
                branch="ALL",
                activo=True
            )
            db.session.add(admin)
        
        # 2. Crear operadores para cada Branch
        sedes = ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"]
        for sede in sedes:
            op_sede = Operador.query.filter(db.func.lower(Operador.dni) == sede.lower()).first()
            if not op_sede:
                pwd_sede = generate_password_hash("Vip@123")
                nuevo_op = Operador(
                    nombre=f"Operador Sede {sede}",
                    dni=sede.upper(),
                    correo=f"{sede.lower()}@botvip.com",
                    password_hash=pwd_sede,
                    rol="operador",
                    branch=sede.upper(),
                    activo=True
                )
                db.session.add(nuevo_op)
                print(f"Semilla creada para sede {sede}")
        
        # 3. Crear/restablecer usuario NOC
        noc = Operador.query.filter(db.func.lower(Operador.dni) == "noc").first()
        if noc:
            noc.dni = "NOC"
            noc.password_hash = generate_password_hash("Bitel@123")
            noc.rol = "noc"
            noc.branch = "ALL"
            noc.correo = "noc@botvip.com"
        else:
            default_pwd = generate_password_hash("Bitel@123")
            noc = Operador(
                nombre="Usuario NOC",
                dni="NOC",
                correo="noc@botvip.com",
                password_hash=default_pwd,
                rol="noc",
                branch="ALL",
                activo=True
            )
            db.session.add(noc)
        
        db.session.commit()
        print("Sedes, Administrador y NOC sembrados correctamente.")
    except Exception as e:
        print("Error al crear operadores por defecto:", e)
        db.session.rollback()


def buscar_y_copiar_materiales_previos(cuenta, target_averia, averias_por_cuenta=None):
    if averias_por_cuenta is not None:
        candidates = averias_por_cuenta.get(cuenta, [])
        valid_previos = []
        for av in candidates:
            if av.estado != "REPARADO":
                continue
            if target_averia.id and av.id == target_averia.id:
                continue
            tiene_mats = (
                (av.material_cable_m or 0) > 0 or
                (av.material_conectores or 0) > 0 or
                (av.material_rosetas or 0) > 0 or
                (av.material_mangas or 0) > 0 or
                (av.material_acopladores or 0) > 0 or
                (av.materiales_json and av.materiales_json != "{}")
            )
            if tiene_mats:
                valid_previos.append(av)
        if valid_previos:
            valid_previos.sort(key=lambda x: x.id, reverse=True)
            previo = valid_previos[0]
            target_averia.materiales_json = previo.materiales_json
            target_averia.material_cable_m = previo.material_cable_m
            target_averia.material_conectores = previo.material_conectores
            target_averia.material_rosetas = previo.material_rosetas
            target_averia.material_mangas = previo.material_mangas
            target_averia.material_acopladores = previo.material_acopladores
            target_averia.tecnico_id = previo.tecnico_id
            target_averia.tipificacion = previo.tipificacion
            comentario_sync = target_averia.material_comentarios or ""
            if previo.material_comentarios:
                target_averia.material_comentarios = f"{previo.material_comentarios} (Sincronizado: {comentario_sync})"
            else:
                target_averia.material_comentarios = comentario_sync
            return True
        return False

    # Buscar una avería resuelta con materiales para la misma cuenta (fallback de consulta directa)
    previo = Averia.query.filter(
        Averia.cuenta == cuenta,
        Averia.estado == "REPARADO",
        Averia.id != target_averia.id if target_averia.id else True,
        (
            (Averia.material_cable_m > 0) |
            (Averia.material_conectores > 0) |
            (Averia.material_rosetas > 0) |
            (Averia.material_mangas > 0) |
            (Averia.material_acopladores > 0) |
            ((Averia.materiales_json != None) & (Averia.materiales_json != "{}"))
        )
    ).order_by(Averia.id.desc()).first()
    
    if previo:
        target_averia.materiales_json = previo.materiales_json
        target_averia.material_cable_m = previo.material_cable_m
        target_averia.material_conectores = previo.material_conectores
        target_averia.material_rosetas = previo.material_rosetas
        target_averia.material_mangas = previo.material_mangas
        target_averia.material_acopladores = previo.material_acopladores
        target_averia.tecnico_id = previo.tecnico_id
        target_averia.tipificacion = previo.tipificacion
        # Preservar el comentario de drive pero adjuntar el anterior
        comentario_sync = target_averia.material_comentarios or ""
        if previo.material_comentarios:
            target_averia.material_comentarios = f"{previo.material_comentarios} (Sincronizado: {comentario_sync})"
        else:
            target_averia.material_comentarios = comentario_sync
        return True
    return False


def es_ticket_no_odn(detalles, status_ont, status_caja, material_comentarios=None):
    det = (detalles or "").lower()
    ont = (status_ont or "").lower()
    caja = (status_caja or "").lower()
    com = (material_comentarios or "").lower()
    
    def contains_word(text, word):
        if word not in text:
            return False
        if word == "online":
            if "not online" in text or "no online" in text or "not_online" in text:
                import re
                occurrences = [m.start() for m in re.finditer("online", text)]
                for start in occurrences:
                    pre_text = text[max(0, start-5):start]
                    if "not" not in pre_text and "no" not in pre_text:
                        return True
                return False
        return True

    if contains_word(det, "online") or contains_word(ont, "online") or contains_word(caja, "online") or contains_word(com, "estado: online"):
        return True
    if "contacto" in det or "contacto" in ont or "contacto" in caja or "estado: contacto" in com:
        return True
    if "desea" in det or "desea" in ont or "desea" in caja or "desea" in com:
        return True
    if "contesta" in det or "contesta" in ont or "contesta" in caja or "contesta" in com:
        return True
    if "baja" in det or "baja" in ont or "baja" in caja or "baja" in com:
        return True
        
    return False


def sincronizar_drive():
    url = "https://docs.google.com/spreadsheets/d/1eaNxCpm8JF1JcZS3_ldwMRINGYFaW6RsQQWybvRi_P8/export?format=csv&gid=1775459558"
    try:
        response = requests.get(url, timeout=15)
        if response.status_code != 200:
            return False, f"Error de conexión con Google Sheets (Status: {response.status_code})"
        
        content = response.content.decode('utf-8-sig')
        f = io.StringIO(content)
        reader = csv.reader(f)
        
        # Leer cabecera
        header = next(reader, None)
        if not header:
            return False, "El archivo de Google Sheets está vacío."
            
        # Detectar errores comunes de fórmulas en el documento
        has_formula_errors = any("#REF!" in str(col) or "#VALUE!" in str(col) or "#N/A" in str(col) for col in header)
        if has_formula_errors:
            return False, "El archivo de Google Sheets contiene errores de fórmula (#REF!, #VALUE!, #N/A) en la cabecera. Por favor, espera a que cargue o verifica el documento."
        
        # Mapeo de columnas por índice
        indices = {}
        for i, col in enumerate(header):
            col_name = col.strip().upper().replace("Ó", "O").replace("Í", "I")
            indices[col_name] = i
        
        # Verificar columnas críticas
        columnas_requeridas = ["BRANCH", "CUENTA", "COORDENADAS"]
        for col in columnas_requeridas:
            if col not in indices:
                columnas_encontradas = [c for c in indices.keys() if c]
                return False, f"Falta la columna requerida: {col}. Columnas encontradas: {columnas_encontradas}"
        
        cuentas_drive = set()
        nuevos = 0
        actualizados = 0
        nuevos_por_branch = {}
        
        # Cargar todas las averías de la base de datos en memoria para evitar consultas en bucle
        all_averias = Averia.query.all()
        averias_por_wo = {}
        averias_por_cuenta = {}
        for av in all_averias:
            if av.codigo_wo:
                if av.codigo_wo not in averias_por_wo:
                    averias_por_wo[av.codigo_wo] = []
                averias_por_wo[av.codigo_wo].append(av)
            if av.cuenta:
                if av.cuenta not in averias_por_cuenta:
                    averias_por_cuenta[av.cuenta] = []
                averias_por_cuenta[av.cuenta].append(av)
        
        for row in reader:
            if not row or len(row) <= max(indices.values()):
                continue
            
            branch = row[indices["BRANCH"]].strip()
            cuenta = row[indices["CUENTA"]].strip()
            
            # Saltar si no hay cuenta o branch
            if not cuenta or not branch:
                continue
            
            cuentas_drive.add(cuenta)
            
            # Extraer campos
            codigo_wo = row[indices.get("CODIGO DE WO")].strip() if "CODIGO DE WO" in indices else ""
            detalles = row[indices.get("DETALLES")].strip() if "DETALLES" in indices else ""
            
            dias_str = row[indices.get("DIAS PENDIENTES")].strip() if "DIAS PENDIENTES" in indices else ""
            try:
                dias_pendientes = float(dias_str) if dias_str else 0.0
            except ValueError:
                dias_pendientes = 0.0
                
            status_ont = row[indices.get("ESTADO")].strip() if "ESTADO" in indices else ""
            status_caja = row[indices.get("STATUS DE LA CAJA")].strip() if "STATUS DE LA CAJA" in indices else ""
            contrata = row[indices.get("CONTRATA")].strip() if "CONTRATA" in indices else ""
            periodo_pendiente = row[indices.get("PERIODO PENDIENTE")].strip() if "PERIODO PENDIENTE" in indices else ""
            site = row[indices.get("SITE")].strip() if "SITE" in indices else ""
            caja = row[indices.get("CAJA")].strip() if "CAJA" in indices else ""
            coordenadas = row[indices.get("COORDENADAS")].strip() if "COORDENADAS" in indices else ""
            
            # Check if this is a non-ODN ticket (ONLINE, contacto, etc.)
            if es_ticket_no_odn(detalles, status_ont, status_caja):
                # Delete any existing tickets under this account that are themselves non-ODN
                existing_tickets = Averia.query.filter_by(cuenta=cuenta).all()
                for av in existing_tickets:
                    status_for_check = av.estado if av.estado == "ONLINE" else ""
                    if es_ticket_no_odn(av.detalles, status_for_check, av.status_caja, av.material_comentarios):
                        db.session.delete(av)
                continue # Skip importing/updating this row
                
            # Verificar si en el Drive figura como resuelto
            status_upper = status_ont.upper().strip()
            status_caja_upper = status_caja.upper().strip()
            resuelto_keywords = ["REPARADO", "SOLUCIONADO", "CERRADO", "OK", "ATENDIDO", "RESUELTO"]
            es_resuelto_drive = any(k in status_upper or k in status_caja_upper for k in resuelto_keywords)

            # Buscar en memoria
            averia = None
            if codigo_wo:
                candidates_wo = averias_por_wo.get(codigo_wo, [])
                candidates_wo_pending = [a for a in candidates_wo if a.estado == "PENDIENTE"]
                if candidates_wo_pending:
                    averia = candidates_wo_pending[0]
                elif candidates_wo:
                    candidates_wo_sorted = sorted(candidates_wo, key=lambda x: x.id, reverse=True)
                    averia = candidates_wo_sorted[0]
            
            if not averia:
                candidates_cuenta = averias_por_cuenta.get(cuenta, [])
                candidates_cuenta_pending = [a for a in candidates_cuenta if a.estado == "PENDIENTE"]
                if candidates_cuenta_pending:
                    candidates_cuenta_pending_sorted = sorted(candidates_cuenta_pending, key=lambda x: x.id, reverse=True)
                    averia = candidates_cuenta_pending_sorted[0]
                else:
                    candidates_cuenta_reparado = [a for a in candidates_cuenta if a.estado == "REPARADO"]
                    if candidates_cuenta_reparado:
                        candidates_cuenta_reparado_sorted = sorted(candidates_cuenta_reparado, key=lambda x: x.id, reverse=True)
                        ultimo_reparado = candidates_cuenta_reparado_sorted[0]
                        if es_resuelto_drive:
                            averia = ultimo_reparado
                        else:
                            if ultimo_reparado.fecha_resolucion:
                                diff_horas = (obtener_hora_peru() - ultimo_reparado.fecha_resolucion).total_seconds() / 3600.0
                                if diff_horas <= 24.0:
                                    averia = ultimo_reparado

            if averia:
                # Si ya existe, actualizar datos del drive solo si está pendiente localmente
                if averia.estado == "PENDIENTE":
                    if es_resuelto_drive:
                        averia.estado = "REPARADO"
                        averia.fecha_resolucion = obtener_hora_peru()
                        averia.material_comentarios = f"Marcado como resuelto en el Drive (Estado: {status_ont or status_caja})"
                        # Copiar materiales si existían anteriormente
                        buscar_y_copiar_materiales_previos(cuenta, averia, averias_por_cuenta)
                    else:
                        averia.branch = branch
                        averia.codigo_wo = codigo_wo
                        averia.detalles = detalles
                        averia.dias_pendientes = dias_pendientes
                        averia.status_caja = status_caja if status_caja else status_ont
                        averia.contrata = contrata
                        averia.periodo_pendiente = periodo_pendiente
                        averia.site = site
                        averia.caja = caja
                        averia.coordenadas = coordenadas
                    actualizados += 1
                elif averia.estado == "REPARADO":
                    # Si ya está reparada en BD pero no por un técnico de la web (no tiene tecnico_id y no tiene materiales)
                    # y el drive ahora dice que no está resuelto (es_resuelto_drive == False), entonces revertir a PENDIENTE
                    tiene_materiales = (
                        (averia.material_cable_m or 0) > 0 or
                        (averia.material_conectores or 0) > 0 or
                        (averia.material_rosetas or 0) > 0 or
                        (averia.material_mangas or 0) > 0 or
                        (averia.material_acopladores or 0) > 0 or
                        (averia.materiales_json and averia.materiales_json != "{}")
                    )
                    # Si no tiene materiales, intentamos copiarlos por si existían anteriormente
                    if not tiene_materiales:
                        buscar_y_copiar_materiales_previos(cuenta, averia, averias_por_cuenta)
                        tiene_materiales = (
                            (averia.material_cable_m or 0) > 0 or
                            (averia.material_conectores or 0) > 0 or
                            (averia.material_rosetas or 0) > 0 or
                            (averia.material_mangas or 0) > 0 or
                            (averia.material_acopladores or 0) > 0 or
                            (averia.materiales_json and averia.materiales_json != "{}")
                        )
                    if not averia.tecnico_id and not tiene_materiales and not es_resuelto_drive:
                        averia.estado = "PENDIENTE"
                        averia.fecha_resolucion = None
                        averia.material_comentarios = None
                        averia.branch = branch
                        averia.codigo_wo = codigo_wo
                        averia.detalles = detalles
                        averia.dias_pendientes = dias_pendientes
                        averia.status_caja = status_caja if status_caja else status_ont
                        averia.contrata = contrata
                        averia.periodo_pendiente = periodo_pendiente
                        averia.site = site
                        averia.caja = caja
                        averia.coordenadas = coordenadas
                        actualizados += 1
            else:
                # Crear nueva avería
                nueva = Averia(
                    branch=branch,
                    codigo_wo=codigo_wo,
                    cuenta=cuenta,
                    detalles=detalles,
                    dias_pendientes=dias_pendientes,
                    estado="REPARADO" if es_resuelto_drive else "PENDIENTE",
                    fecha_resolucion=obtener_hora_peru() if es_resuelto_drive else None,
                    material_comentarios=f"Marcado como resuelto en el Drive al importar" if es_resuelto_drive else None,
                    status_caja=status_caja if status_caja else status_ont,
                    contrata=contrata,
                    periodo_pendiente=periodo_pendiente,
                    site=site,
                    caja=caja,
                    coordenadas=coordenadas,
                    origen="SHEETS",
                    fecha_creacion=obtener_hora_peru()
                )
                if es_resuelto_drive:
                    buscar_y_copiar_materiales_previos(cuenta, nueva, averias_por_cuenta)
                db.session.add(nueva)
                if cuenta not in averias_por_cuenta:
                    averias_por_cuenta[cuenta] = []
                averias_por_cuenta[cuenta].append(nueva)
                nuevos += 1
                if branch not in nuevos_por_branch:
                    nuevos_por_branch[branch] = []
                nuevos_por_branch[branch].append(nueva)
                
        # Obtener todas las averías pendientes originadas en el Drive y filtrarlas en Python para evitar consultas SQL lentas con miles de parámetros
        sheets_pending = Averia.query.filter_by(origen="SHEETS", estado="PENDIENTE").all()
        cerrados_ext = [av for av in sheets_pending if av.cuenta not in cuentas_drive]
        
        for av in cerrados_ext:
            av.estado = "REPARADO"
            av.fecha_resolucion = obtener_hora_peru()
            av.material_comentarios = "Cerrado automáticamente al no figurar en la lista del Drive."
        
        db.session.commit()
        total_cerrados = len(cerrados_ext)
        
        # Sincronizar sites y boxes en segundo plano para evitar timeout de la solicitud HTTP
        import threading
        def sync_bg():
            with app.app_context():
                try:
                    sincronizar_sites()
                    from models import Box
                    if db.session.query(Box).first() is None:
                        print("Sincronizando boxes en segundo plano...")
                        sincronizar_boxes()
                    else:
                        print("Sincronización de boxes omitida porque ya existen registros.")
                except Exception as e:
                    print("Error en sincronización en segundo plano de sites/boxes:", e)
                finally:
                    db.session.remove()
        
        threading.Thread(target=sync_bg).start()
            
        return True, f"Sincronización exitosa: {nuevos} creados, {actualizados} actualizados y {total_cerrados} cerrados. (La sincronización de cajas y nodos continúa en segundo plano)."
    except Exception as e:
        db.session.rollback()
        print("Error en sincronización de drive:", e)
        return False, f"Error en sincronización: {str(e)}"


@app.route("/diagnostico_materiales")
def diagnostico_materiales():
    try:
        from sqlalchemy import func
        reparados = Averia.query.filter_by(estado="REPARADO").order_by(Averia.id.desc()).all()
        
        # Ejecutar la migración manualmente y capturar logs
        migrated_log = []
        
        # Buscar todas las averías REPARADO que no tienen materiales
        reparadas_sin_mats = Averia.query.filter(
            Averia.estado == "REPARADO",
            (Averia.material_cable_m.is_(None) | (Averia.material_cable_m == 0)),
            (Averia.material_conectores.is_(None) | (Averia.material_conectores == 0)),
            (Averia.material_rosetas.is_(None) | (Averia.material_rosetas == 0)),
            (Averia.material_mangas.is_(None) | (Averia.material_mangas == 0)),
            (Averia.material_acopladores.is_(None) | (Averia.material_acopladores == 0)),
            (Averia.materiales_json.is_(None) | (Averia.materiales_json == "{}") | (Averia.materiales_json == ""))
        ).all()
        
        migrated_count = 0
        for av in reparadas_sin_mats:
            # Buscar previos
            previo = Averia.query.filter(
                Averia.cuenta == av.cuenta,
                Averia.estado == "REPARADO",
                Averia.id != av.id,
                (
                    (Averia.material_cable_m > 0) |
                    (Averia.material_conectores > 0) |
                    (Averia.material_rosetas > 0) |
                    (Averia.material_mangas > 0) |
                    (Averia.material_acopladores > 0) |
                    ((Averia.materiales_json != None) & (Averia.materiales_json != "{}"))
                )
            ).order_by(Averia.id.desc()).first()
            
            if previo:
                av.materiales_json = previo.materiales_json
                av.material_cable_m = previo.material_cable_m
                av.material_conectores = previo.material_conectores
                av.material_rosetas = previo.material_rosetas
                av.material_mangas = previo.material_mangas
                av.material_acopladores = previo.material_acopladores
                av.tecnico_id = previo.tecnico_id
                av.tipificacion = previo.tipificacion
                comentario_sync = av.material_comentarios or ""
                if previo.material_comentarios:
                    av.material_comentarios = f"{previo.material_comentarios} (Sincronizado: {comentario_sync})"
                else:
                    av.material_comentarios = comentario_sync
                migrated_log.append(f"Restaurada cuenta {av.cuenta} (ID {av.id}) desde ID {previo.id} con Cable={previo.material_cable_m}, Conectores={previo.material_conectores}")
                migrated_count += 1
            else:
                otros = Averia.query.filter(Averia.cuenta == av.cuenta, Averia.id != av.id).all()
                otros_ids = [str(o.id) for o in otros]
                migrated_log.append(f"Sin restaurar {av.cuenta} (ID {av.id}): No hay previos con materiales. Otros tickets de la cuenta en BD: [{', '.join(otros_ids)}]")
                
        if migrated_count > 0:
            db.session.commit()
            
        html = f"""
        <html>
        <head><title>Diagnostico de Materiales</title></head>
        <body style="font-family: Arial, sans-serif; padding: 20px; line-height: 1.6;">
            <h2>Diagnostico de Base de Datos - Materiales Perdidos</h2>
            <p><strong>Total tickets REPARADO en BD:</strong> {len(reparados)}</p>
            <p><strong>Total tickets REPARADO sin materiales:</strong> {len(reparadas_sin_mats)}</p>
            <p><strong>Tickets restaurados en esta pasada:</strong> {migrated_count}</p>
            
            <h3>Logs de Restauración:</h3>
            <pre style="background: #f4f4f4; padding: 15px; border: 1px solid #ccc; max-height: 400px; overflow-y: scroll;">
{"\n".join(migrated_log)}
            </pre>
            
            <h3>Últimos 30 tickets REPARADO en el sistema:</h3>
            <table border="1" cellpadding="5" style="border-collapse: collapse; width: 100%;">
                <tr style="background: #eee;">
                    <th>ID</th>
                    <th>Cuenta</th>
                    <th>Estado</th>
                    <th>Origen</th>
                    <th>Técnico ID</th>
                    <th>Cable (m)</th>
                    <th>Conectores</th>
                    <th>Mufas</th>
                    <th>Rosetas</th>
                    <th>Acopladores</th>
                    <th>Agrupados</th>
                    <th>Comentarios</th>
                </tr>
        """
        for av in reparados[:30]:
            html += f"""
                <tr>
                    <td>{av.id}</td>
                    <td>{av.cuenta}</td>
                    <td>{av.estado}</td>
                    <td>{av.origen}</td>
                    <td>{av.tecnico_id}</td>
                    <td>{av.material_cable_m}</td>
                    <td>{av.material_conectores}</td>
                    <td>{av.material_mangas}</td>
                    <td>{av.material_rosetas}</td>
                    <td>{av.material_acopladores}</td>
                    <td>{av.cuentas_asociadas}</td>
                    <td>{av.material_comentarios}</td>
                </tr>
            """
        html += """
            </table>
        </body>
        </html>
        """
        return html
    except Exception as e:
        return f"Error en diagnóstico: {str(e)}"


@app.route("/debug_ticket")
def debug_ticket():
    from sqlalchemy import func
    averias = Averia.query.filter(Averia.cuenta.ilike('%19061621%')).all()
    
    # Run the exact same sums query
    query_sums = db.session.query(
        func.sum(Averia.material_cable_m).label("cable"),
        func.sum(Averia.material_conectores).label("conectores"),
        func.sum(Averia.material_rosetas).label("rosetas"),
        func.sum(Averia.material_mangas).label("mangas"),
        func.sum(Averia.material_acopladores).label("acopladores")
    ).filter_by(estado="REPARADO")
    
    # Run globally and filtered by JUN
    sums_all = query_sums.first()
    sums_jun = query_sums.filter_by(branch="JUN").first()
    
    results = []
    for av in averias:
        results.append({
            "id": av.id,
            "cuenta": av.cuenta,
            "estado": av.estado,
            "branch": av.branch,
            "coordenadas": av.coordenadas,
            "materiales_json": av.materiales_json,
            "material_acopladores": av.material_acopladores,
            "material_cable_m": av.material_cable_m,
            "material_conectores": av.material_conectores,
            "material_rosetas": av.material_rosetas,
            "material_mangas": av.material_mangas
        })
        
    return jsonify({
        "tickets": results,
        "sums_all": {
            "cable": sums_all.cable if sums_all else None,
            "conectores": sums_all.conectores if sums_all else None,
            "rosetas": sums_all.rosetas if sums_all else None,
            "mangas": sums_all.mangas if sums_all else None,
            "acopladores": sums_all.acopladores if sums_all else None
        },
        "sums_jun": {
            "cable": sums_jun.cable if sums_jun else None,
            "conectores": sums_jun.conectores if sums_jun else None,
            "rosetas": sums_jun.rosetas if sums_jun else None,
            "mangas": sums_jun.mangas if sums_jun else None,
            "acopladores": sums_jun.acopladores if sums_jun else None
        }
    })


@app.route("/debug_map_json")
def debug_map_json():
    averias = Averia.query.filter(Averia.cuenta.ilike('%19061621%')).all()
    results = []
    for av in averias:
        lat = None
        lng = None
        if av.coordenadas:
            coords = av.coordenadas.split(",")
            if len(coords) == 2:
                try:
                    lat = float(coords[0].strip())
                    lng = float(coords[1].strip())
                except ValueError:
                    pass
        results.append({
            "id": av.id,
            "branch": av.branch,
            "cuenta": av.cuenta,
            "estado": av.estado,
            "lat": lat,
            "lng": lng,
            "cable": av.material_cable_m or 0,
            "conectores": av.material_conectores or 0,
            "rosetas": av.material_rosetas or 0,
            "mangas": av.material_mangas or 0,
            "acopladores": av.material_acopladores or 0
        })
    return jsonify(results)


@app.route("/manifest.json")
def serve_manifest():
    return app.send_static_file("manifest.json")


@app.route("/sw.js")
def serve_sw():
    response = app.send_static_file("sw.js")
    response.headers["Service-Worker-Allowed"] = "/"
    return response


@app.route("/api/caja_coordenadas", methods=["GET"])
@login_requerido
def get_caja_coordenadas():
    caja_code = request.args.get("caja", "").strip().upper()
    if not caja_code:
        return jsonify({"success": False, "message": "No se especificó el código de caja"}), 400
        
    box = Box.query.filter_by(caja=caja_code).first()
    if box:
        return jsonify({
            "success": True,
            "caja": box.caja,
            "lat": box.latitud,
            "lng": box.longitud
        })
    else:
        return jsonify({"success": False, "message": "Coordenadas no encontradas para la caja especificada"}), 404


@app.route("/api/averias/alertas", methods=["GET"])
@login_requerido
def api_alertas_averias():
    last_id = request.args.get("last_id", type=int)
    if last_id is None:
        return {"alertas": []}, 200
        
    branch = session.get("operador_branch")
    es_admin = session.get("operador_rol") == "admin"
    
    query = Averia.query.filter(Averia.estado == "PENDIENTE", Averia.id > last_id)
    if not es_admin and branch != "ALL":
        query = query.filter_by(branch=branch)
        
    nuevas = query.order_by(Averia.id.asc()).all()
    
    alertas = []
    for av in nuevas:
        alertas.append({
            "id": av.id,
            "cuenta": av.cuenta,
            "branch": av.branch,
            "caja": av.caja or "N/A",
            "detalles": av.detalles or "Sin descripción",
            "coordenadas": av.coordenadas or ""
        })
        
    return {"alertas": alertas}, 200


@app.route("/diagnostico-operadores")
def diagnostico_operadores():
    try:
        crear_operador_defecto()
        ops = Operador.query.all()
        resultado = f"<h3>Diagnóstico de Operadores y Sedes</h3>"
        resultado += "<table border='1' cellpadding='5' style='border-collapse:collapse;'>"
        resultado += "<tr><th>ID</th><th>Nombre</th><th>DNI</th><th>Sede</th><th>Rol</th><th>Activo</th><th>Verificación Password (Bitel@123 / Vip@123)</th></tr>"
        
        for op in ops:
            matches_pwd = check_password_hash(op.password_hash, "Bitel@123") or check_password_hash(op.password_hash, "Vip@123")
            resultado += f"<tr>"
            resultado += f"<td>{op.id}</td>"
            resultado += f"<td>{op.nombre}</td>"
            resultado += f"<td>{op.dni}</td>"
            resultado += f"<td>{op.branch}</td>"
            resultado += f"<td>{op.rol}</td>"
            resultado += f"<td>{op.activo}</td>"
            resultado += f"<td>{'CORRECTA' if matches_pwd else 'OTRA CONTRASEÑA'}</td>"
            resultado += f"</tr>"
            
        resultado += "</table>"
        resultado += "<br><a href='/login'>Ir al Login</a>"
        return resultado
    except Exception as e:
        db.session.rollback()
        return f"Error en diagnóstico: {str(e)}"


@app.route("/login", methods=["GET", "POST"])
def login():
    if "operador_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        dni = request.form["dni"].strip()
        password = request.form["password"].strip()

        operador = Operador.query.filter(db.func.lower(Operador.dni) == db.func.lower(dni), Operador.activo == True).first()
        if operador and check_password_hash(operador.password_hash, password):
            session["operador_id"] = operador.id
            session["operador_nombre"] = operador.nombre
            session["operador_dni"] = operador.dni
            session["operador_rol"] = operador.rol
            session["operador_branch"] = operador.branch
            
            flash(f"¡Bienvenido de nuevo, {operador.nombre} ({operador.branch})!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("DNI/Usuario o contraseña incorrectos.", "danger")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if "operador_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        dni = request.form["dni"].strip()
        nombre = request.form["nombre"].strip()
        correo = request.form["correo"].strip()
        password = request.form["password"].strip()
        branch = request.form["branch"].strip().upper()

        existente = Operador.query.filter((db.func.lower(Operador.dni) == db.func.lower(dni)) | (db.func.lower(Operador.correo) == db.func.lower(correo))).first()
        if existente:
            flash("Ya existe un operador con ese DNI o correo.", "danger")
            return render_template("register.html")

        pwd_hash = generate_password_hash(password or dni)

        nuevo = Operador(
            dni=dni,
            nombre=nombre,
            correo=correo,
            password_hash=pwd_hash,
            rol="operador",
            branch=branch,
            activo=True
        )
        db.session.add(nuevo)
        db.session.commit()
        flash("Cuenta registrada con éxito. Ya puedes iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("register.html")


@app.route("/operadores")
@login_requerido
def listar_operadores():
    if session.get("operador_rol") != "admin":
        flash("Acceso denegado: Se requieren permisos de administrador.", "danger")
        return redirect(url_for("dashboard"))

    operadores = Operador.query.order_by(Operador.id.desc()).all()
    stats = obtener_estadisticas(branch=session.get("operador_branch"), es_admin=True)
    return render_template("operadores.html", operadores=operadores, stats=stats)



@app.route("/logout")
def logout():
    session.clear()
    flash("Has cerrado sesión correctamente.", "info")
    return redirect(url_for("login"))


@app.route("/cambiar-password", methods=["POST"])
@login_requerido
def cambiar_password():
    password_actual = request.form["password_actual"].strip()
    nueva_password = request.form["nueva_password"].strip()
    confirmar_password = request.form["confirmar_password"].strip()

    if nueva_password != confirmar_password:
        flash("La nueva contraseña y la confirmación no coinciden.", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    operador = db.session.get(Operador, session["operador_id"])
    if operador and check_password_hash(operador.password_hash, password_actual):
        operador.password_hash = generate_password_hash(nueva_password)
        db.session.commit()
        flash("Tu contraseña ha sido cambiada exitosamente.", "success")
    else:
        flash("La contraseña actual es incorrecta.", "danger")

    return redirect(request.referrer or url_for("dashboard"))


@app.route("/plan-diario", methods=["GET"])
@login_requerido
def generar_plan_diario():
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    user_branch = session.get("operador_branch")
    
    branch_arg = request.args.get("branch")
    if not es_admin and not es_noc and user_branch != "ALL":
        target_branch = user_branch
    else:
        target_branch = branch_arg if branch_arg else (user_branch if user_branch != "ALL" else "LI1")
        
    target_branch = target_branch.strip().upper()
    
    pending_tickets = Averia.query.filter_by(branch=target_branch, estado="PENDIENTE").all()
    
    def parse_caja_local(caja_str, site_val):
        if not caja_str:
            return site_val or "", "", "", ""
        parts = [p.strip().upper() for p in caja_str.split("-") if p.strip()]
        site = site_val or ""
        xbox = ""
        hubox = ""
        subox = ""
        if len(parts) > 0:
            site = parts[0]
        for p in parts[1:]:
            if p.startswith("XB"):
                xbox = p
            elif p.startswith("HB"):
                hubox = p
            else:
                subox = p
        return site, xbox, hubox, subox

    tickets_data = []
    for t in pending_tickets:
        site_p, xbox_p, hubox_p, subox_p = parse_caja_local(t.caja, t.site)
        tickets_data.append({
            "id": t.id,
            "site": t.site or "SIN_SITE",
            "caja": t.caja or "SIN_CAJA",
            "xbox": xbox_p or "",
            "hubox": hubox_p or "",
            "detalles": t.detalles or "Sin detalles",
            "cuenta": t.cuenta or ""
        })
        
    operators = Operador.query.filter_by(branch=target_branch, activo=True).order_by(Operador.nombre).all()
    if not operators:
        operators = Operador.query.filter_by(activo=True).order_by(Operador.nombre).all()
        
    sedes = ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"]
    
    return render_template(
        "plan_diario.html",
        tickets_data=tickets_data,
        operators=operators,
        target_branch=target_branch,
        sedes=sedes
    )


@app.route("/", methods=["GET"])
@login_requerido
def dashboard():
    if session.get("operador_rol") == "noc":
        return redirect(url_for("noc_dashboard"))
        
    es_admin = session.get("operador_rol") == "admin"
    branch = session.get("operador_branch")
    
    stats = obtener_estadisticas(branch=branch, es_admin=es_admin)
    
    # Obtener las averías pendientes y las reparadas en los últimos 12 meses
    from datetime import datetime, timedelta
    hace_12_meses = obtener_hora_peru() - timedelta(days=365)
    query_pendientes = Averia.query.filter_by(estado="PENDIENTE")
    query_reparadas = Averia.query.filter(Averia.estado == "REPARADO", Averia.fecha_resolucion >= hace_12_meses)
    
    if not es_admin and branch != "ALL":
        query_pendientes = query_pendientes.filter_by(branch=branch)
        query_reparadas = query_reparadas.filter_by(branch=branch)
        
    averias_pendientes = query_pendientes.order_by(Averia.dias_pendientes.desc().nullslast()).all()
    averias_reparadas = query_reparadas.order_by(Averia.fecha_resolucion.desc()).all()
    averias_totales = averias_pendientes + averias_reparadas
    
    # Serializar datos para Leaflet map
    map_data = []
    for av in averias_totales:
        lat = None
        lng = None
        if av.coordenadas:
            coords = av.coordenadas.split(",")
            if len(coords) == 2:
                try:
                    lat = float(coords[0].strip())
                    lng = float(coords[1].strip())
                except ValueError:
                    pass
        map_data.append({
            "id": av.id,
            "branch": av.branch,
            "cuenta": av.cuenta or "Sin cuenta",
            "codigo_wo": av.codigo_wo or "N/A",
            "detalles": av.detalles or "Sin descripción",
            "contrata": av.contrata or "Sin contrata",
            "site": av.site or "N/A",
            "caja": av.caja or "N/A",
            "dias": av.dias_pendientes or 0,
            "estado": av.estado,
            "lat": lat,
            "lng": lng,
            "materiales_json": av.materiales_json or "{}",
            "comentarios": av.material_comentarios or "",
            "tipificacion": av.tipificacion or "",
            "fecha_creacion": av.fecha_creacion.strftime("%Y-%m-%d") if av.fecha_creacion else "",
            "fecha_resolucion": av.fecha_resolucion.strftime("%Y-%m-%d") if av.fecha_resolucion else "",
            "cable": av.material_cable_m or 0,
            "conectores": av.material_conectores or 0,
            "rosetas": av.material_rosetas or 0,
            "mangas": av.material_mangas or 0,
            "acopladores": av.material_acopladores or 0
        })
                    
    # Habilitar formulario manual para todos
    es_provincia = True
    
    from sqlalchemy import func
    # Calculate material totals for this branch/sede
    query_sums = db.session.query(
        func.sum(Averia.material_cable_m).label("cable"),
        func.sum(Averia.material_conectores).label("conectores"),
        func.sum(Averia.material_rosetas).label("rosetas"),
        func.sum(Averia.material_mangas).label("mangas"),
        func.sum(Averia.material_acopladores).label("acopladores")
    ).filter_by(estado="REPARADO")
    
    if not es_admin and branch != "ALL":
        query_sums = query_sums.filter_by(branch=branch)
        
    sums = query_sums.first()
    
    materials_totals = {
        "cable": (sums.cable if sums and sums.cable is not None else 0),
        "conectores": (sums.conectores if sums and sums.conectores is not None else 0),
        "rosetas": (sums.rosetas if sums and sums.rosetas is not None else 0),
        "mangas": (sums.mangas if sums and sums.mangas is not None else 0),
        "acopladores": (sums.acopladores if sums and sums.acopladores is not None else 0)
    }
    
    import json
    return render_template(
        "dashboard.html", 
        stats=stats, 
        map_json=json.dumps(map_data),
        averias_list=averias_totales,
        es_provincia=es_provincia,
        materials=materials_totals
    )


@app.route("/averias/sync", methods=["POST", "GET"])
@login_requerido
def sync_sheets():
    # Solo permitir sincronizar a Lima y Administradores
    branch = session.get("operador_branch")
    if branch not in ["LI1", "LI2", "LI3", "LI4", "LI7", "ALL"] and session.get("operador_rol") != "admin":
        flash("Tu sede no tiene habilitada la sincronización automática.", "warning")
        return redirect(url_for("dashboard"))
        
    success, message = sincronizar_drive()
    if success:
        flash(message, "success")
    else:
        flash(message, "danger")
    return redirect(url_for("dashboard"))


def ajustar_stock_por_consumo(branch, old_mats_dict, new_mats_dict):
    """
    Compara old_mats_dict y new_mats_dict, calcula la diferencia (new_qty - old_qty)
    y descuenta esa diferencia de StockBranch.stock_actual para la sede dada.
    """
    all_keys = set(list(old_mats_dict.keys()) + list(new_mats_dict.keys()))
    for key in all_keys:
        parts = key.split("|")
        if len(parts) < 2:
            continue
        codigo = parts[0]
        nombre = parts[1]
        
        old_qty = old_mats_dict.get(key, 0)
        new_qty = new_mats_dict.get(key, 0)
        
        try:
            old_qty = int(old_qty)
        except (ValueError, TypeError):
            old_qty = 0
            
        try:
            new_qty = int(new_qty)
        except (ValueError, TypeError):
            new_qty = 0
            
        diff = new_qty - old_qty
        if diff == 0:
            continue
            
        reg = StockBranch.query.filter_by(branch=branch, material_nombre=nombre).first()
        if not reg:
            reg = StockBranch(
                branch=branch,
                material_codigo=codigo,
                material_nombre=nombre,
                stock_actual=0
            )
            db.session.add(reg)
            
        reg.stock_actual = max(0, reg.stock_actual - diff)


@app.route("/averias/resolver/<int:id>", methods=["POST"])
@login_requerido
def resolver_averia(id):
    if session.get("operador_rol") == "noc":
        flash("No tienes permisos para realizar esta acción.", "danger")
        return redirect(url_for("dashboard"))
        
    averia = db.session.get(Averia, id)
    if not averia:
        flash("La avería no existe.", "danger")
        return redirect(url_for("dashboard"))
        
    # Verificar pertenencia a la sede
    branch = session.get("operador_branch")
    if session.get("operador_rol") != "admin" and branch != "ALL" and averia.branch != branch:
        flash("No tienes permisos para resolver averías de otra sede.", "danger")
        return redirect(url_for("dashboard"))
        
    try:
        import json
        materiales_json_str = request.form.get("materiales_json", "{}")
        comentarios = request.form.get("comentarios", "").strip()
        
        # Obtener materiales viejos antes de actualizar
        old_mats = averia.materiales_dict
        
        try:
            mats = json.loads(materiales_json_str)
        except Exception:
            mats = {}
            
        # Ajustar el stock según el consumo
        ajustar_stock_por_consumo(averia.branch, old_mats, mats)
        
        averia.estado = "REPARADO"
        averia.fecha_resolucion = obtener_hora_peru()
        averia.tecnico_id = session.get("operador_id")
        averia.materiales_json = materiales_json_str
        averia.material_comentarios = comentarios or "Reparado desde el portal"
        averia.tipificacion = request.form.get("tipificacion", "").strip()
        
        # Calcular valores compatibles para las 5 columnas básicas
        cable_m = 0
        conectores = 0
        rosetas = 0
        mangas = 0
        acopladores = 0
        
        for key, cant in mats.items():
            parts = key.split("|")
            codigo = parts[0]
            nombre = parts[1] if len(parts) > 1 else ""
            
            m_cod, m_nom, m_sec = obtener_material_mapeado(codigo, nombre)
            
            if m_nom == "Cable Drop":
                cable_m += cant
            elif m_nom == "FAC":
                conectores += cant
            elif m_nom == "Waterproof":
                rosetas += cant
            elif m_nom == "Mufas":
                mangas += cant
            elif m_nom == "Preconectorizado":
                acopladores += cant
                
        averia.material_cable_m = cable_m
        averia.material_conectores = conectores
        averia.material_rosetas = rosetas
        averia.material_mangas = mangas
        averia.material_acopladores = acopladores
        
        db.session.commit()
        flash(f"Avería de cuenta {averia.cuenta} resuelta y materiales registrados.", "success")
        if averia.branch in ["ARE", "JUN", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS"]:
            return redirect(url_for("dashboard"))
        return redirect(url_for("agrupar_clientes_averia", id=averia.id))
    except Exception as e:
        db.session.rollback()
        flash(f"Error al resolver avería: {str(e)}", "danger")
        return redirect(url_for("dashboard"))



@app.route("/averias/agrupar/<int:id>", methods=["GET", "POST"])
@login_requerido
def agrupar_clientes_averia(id):
    if session.get("operador_rol") == "noc":
        flash("No tienes permisos para realizar esta acción.", "danger")
        return redirect(url_for("dashboard"))
        
    averia = db.session.get(Averia, id)
    if not averia:
        flash("La avería no existe.", "danger")
        return redirect(url_for("dashboard"))
        
    if averia.branch in ["ARE", "JUN", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS"]:
        flash("La agrupación de averías no está permitida para esta sede.", "warning")
        return redirect(url_for("dashboard"))
        
    # parse caja values
    def parse_caja(caja_str, site_val):
        if not caja_str:
            return site_val or "", "", "", ""
        parts = [p.strip().upper() for p in caja_str.split("-") if p.strip()]
        site = site_val or ""
        xbox = ""
        hubox = ""
        subox = ""
        
        if len(parts) > 0:
            site = parts[0]
            
        for p in parts[1:]:
            if p.startswith("XB"):
                xbox = p
            elif p.startswith("HB"):
                hubox = p
            else:
                subox = p
                
        return site, xbox, hubox, subox
        
    site, xbox, hubox, subox = parse_caja(averia.caja, averia.site)
    
    # 1. Si la avería (principal) está REPARADA y pasaron más de 30 días desde su resolución, bloquear la agrupación
    if averia.estado == "REPARADO" and averia.fecha_resolucion:
        now = datetime.now()
        ref_date = averia.fecha_resolucion.replace(tzinfo=None) if averia.fecha_resolucion.tzinfo else averia.fecha_resolucion
        if (now - ref_date).days > 30:
            flash(f"No hay ninguna avería reparada (principal) en el SITE '{site}' para agrupar. Por favor resuelve el ticket principal primero.", "warning")
            return redirect(url_for("dashboard"))
            

    if request.method == "POST":
        new_xbox = request.form.get("xbox", "").strip().upper()
        new_hubox = request.form.get("hubox", "").strip().upper()
        new_subox = request.form.get("subox", "").strip().upper()
        
        caja_parts = [averia.site or site]
        if new_xbox:
            caja_parts.append(new_xbox)
        if new_hubox:
            caja_parts.append(new_hubox)
        if new_subox:
            caja_parts.append(new_subox)
            
        caja_compuesta = "-".join(caja_parts)
        averia.caja = caja_compuesta
        
        selected_accounts = request.form.getlist("selected_clientes")
        new_set = set(selected_accounts)
        
        # Parse old associated list
        old_list = [c.strip() for c in (averia.cuentas_asociadas or "").split(",") if c.strip()]
        old_set = set(old_list)
        
        # Accounts to add to group (new - old)
        to_add = new_set - old_set
        # Accounts to remove from group (old - new)
        to_remove = old_set - new_set
        
        from sqlalchemy import func
        site_clean = site.strip().upper() if site else ""
        
        # For new ones, mark as REPARADO
        if to_add:
            averias_to_add = Averia.query.filter(
                func.upper(func.trim(Averia.site)) == site_clean,
                Averia.cuenta.in_(to_add),
                Averia.id != averia.id
            ).all()
            for av_g in averias_to_add:
                av_g.estado = "REPARADO"
                av_g.fecha_resolucion = obtener_hora_peru()
                av_g.tecnico_id = session.get("operador_id")
                princ_comm = f" - {averia.material_comentarios}" if averia.material_comentarios else ""
                av_g.material_comentarios = f"Agrupado en la avería principal ({averia.cuenta}){princ_comm}"
                if not averia.tipificacion and av_g.tipificacion:
                    averia.tipificacion = av_g.tipificacion
                if averia.tipificacion:
                    av_g.tipificacion = averia.tipificacion
                av_g.material_cable_m = 0
                av_g.material_conectores = 0
                av_g.material_rosetas = 0
                av_g.material_mangas = 0
                av_g.material_acopladores = 0
                av_g.materiales_json = "{}"
                
        # For removed ones, revert back to PENDING
        if to_remove:
            averias_to_remove = Averia.query.filter(
                func.upper(func.trim(Averia.site)) == site_clean,
                Averia.cuenta.in_(to_remove),
                Averia.id != averia.id
            ).all()
            for av_g in averias_to_remove:
                av_g.estado = "PENDIENTE"
                av_g.fecha_resolucion = None
                av_g.tecnico_id = None
                av_g.material_comentarios = None
                av_g.tipificacion = None
                
        averia.cuentas_asociadas = ", ".join(selected_accounts) if selected_accounts else None
        
        try:
            db.session.commit()
            flash("Asociación de clientes y caja guardada con éxito.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al guardar agrupación: {str(e)}", "danger")
            
        return redirect(url_for("dashboard"))
        
    # GET: query accounts on the same site
    from sqlalchemy import func
    site_clean = site.strip().upper() if site else ""
    associated_list = [c.strip() for c in (averia.cuentas_asociadas or "").split(",") if c.strip()]
    
    # Query all other averias on the same site
    cuentas_query = Averia.query.filter(
        func.upper(func.trim(Averia.site)) == site_clean,
        Averia.id != averia.id
    ).all()
    
    clientes_dict = {c.codigo_cliente: c for c in Cliente.query.all()}
    
    clientes_del_site = []
    vistas = set()
    for row in cuentas_query:
        if row.cuenta and row.cuenta not in vistas:
            is_associated = row.cuenta in associated_list
            
            # Check if this repaired account is already grouped to another main ticket
            is_already_grouped = False
            if row.estado == "REPARADO" and not is_associated:
                if row.cuentas_asociadas:
                    # It is a main ticket of another group
                    is_already_grouped = True
                elif row.material_comentarios and "agrupado" in row.material_comentarios.lower() and "principal" in row.material_comentarios.lower():
                    if f"({averia.cuenta})" not in row.material_comentarios and f"cuenta {averia.cuenta}" not in row.material_comentarios.lower() and f"id {averia.id}" not in row.material_comentarios.lower():
                        # It belongs to another group
                        is_already_grouped = True
            
            # 30 days pending window check (30 days more or 30 days less than the main ticket)
            dias_diff = abs((row.dias_pendientes or 0.0) - (averia.dias_pendientes or 0.0))
            matches_days_window = dias_diff <= 30
            
            # Show if already associated, or if it meets the days window, is not grouped to another ticket and is PENDING/REPARADO
            if is_associated or (matches_days_window and not is_already_grouped and (row.estado == "PENDIENTE" or row.estado == "REPARADO")):
                vistas.add(row.cuenta)
                cl = clientes_dict.get(row.cuenta)
                clientes_del_site.append({
                    "cuenta": row.cuenta,
                    "caja": row.caja or "",
                    "estado": row.estado,
                    "id": row.id,
                    "nombre": cl.nombre if cl else "Cliente de Sheet",
                    "seleccionado": is_associated,
                    "detalles": row.detalles or ""
                })
            
    # Format materials list for principal ticket
    materiales_principal = []
    for key, val in averia.materiales_dict.items():
        if val and val > 0:
            parts = key.split("|")
            name = parts[1] if len(parts) > 1 else parts[0]
            materiales_principal.append(f"{val} {name}")
            
    return render_template(
        "agrupar.html",
        averia=averia,
        site=site,
        xbox=xbox,
        hubox=hubox,
        subox=subox,
        clientes_del_site=clientes_del_site,
        materiales_principal=materiales_principal
    )


@app.route("/averias/eliminar/<int:id>", methods=["POST"])
@login_requerido
def eliminar_averia(id):
    es_fbb = (session.get("operador_nombre", "").upper() == "FBB" or 
              session.get("operador_dni", "").upper() == "FBB")
    if session.get("operador_rol") != "admin" or not es_fbb:
        flash("No tienes permisos para eliminar registros. Solo la cuenta de FBB (Admin) puede realizar esta acción.", "danger")
        return redirect(url_for("dashboard"))
        
    averia = Averia.query.get_or_404(id)
    if averia.estado == "REPARADO" and averia.materiales_json:
        try:
            import json
            mats = json.loads(averia.materiales_json)
            # Devolvemos el stock: pasar mats como old_mats y un dict vacio como new_mats
            ajustar_stock_por_consumo(averia.branch, mats, {})
        except Exception as e:
            print("Error devolviendo stock al eliminar averia:", e)
            
    try:
        db.session.delete(averia)
        db.session.commit()
        flash(f"Avería ID {id} ({averia.cuenta or 'Sin Cuenta'}) eliminada correctamente y stock devuelto.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al eliminar la avería: {str(e)}", "danger")
        
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/averias/deshacer/<int:id>", methods=["POST"])
@login_requerido
def deshacer_resolucion_averia(id):
    es_fbb = (session.get("operador_nombre", "").upper() == "FBB" or 
              session.get("operador_dni", "").upper() == "FBB")
    if session.get("operador_rol") != "admin" or not es_fbb:
        flash("No tienes permisos para deshacer resoluciones. Solo la cuenta de FBB (Admin) puede realizar esta acción.", "danger")
        return redirect(url_for("dashboard"))
        
    averia = Averia.query.get_or_404(id)
    if averia.estado != "REPARADO":
        flash("Esta avería ya se encuentra pendiente.", "warning")
        return redirect(request.referrer or url_for("dashboard"))
        
    if averia.materiales_json:
        try:
            import json
            mats = json.loads(averia.materiales_json)
            # Reintegrar materiales al stock
            ajustar_stock_por_consumo(averia.branch, mats, {})
        except Exception as e:
            print("Error devolviendo stock al deshacer resolucion:", e)
            
    try:
        averia.estado = "PENDIENTE"
        averia.fecha_resolucion = None
        averia.material_cable_m = 0
        averia.material_conectores = 0
        averia.material_rosetas = 0
        averia.material_mangas = 0
        averia.material_acopladores = 0
        averia.materiales_json = "{}"
        averia.tipificacion = None
        averia.material_comentarios = None
        
        db.session.commit()
        flash(f"Resolución de avería ID {id} deshecha correctamente. El ticket vuelve a estar PENDIENTE y se reintegró el stock.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al deshacer la resolución: {str(e)}", "danger")
        
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/averias/crear", methods=["POST"])
@login_requerido
def crear_averia_manual():
    branch = session.get("operador_branch")
    es_admin = session.get("operador_rol") == "admin"
    
    try:
        cuenta = request.form.get("cuenta", "").strip()
        accion = request.form.get("accion", "").strip()
        if not cuenta:
            now_str = obtener_hora_peru().strftime("%d%m%H%M")
            cuenta = f"AVERÍA_ODN_{now_str}"
            
        site = request.form["site"].strip().upper()
        xbox = request.form.get("xbox", "").strip().upper()
        hubox = request.form.get("hubox", "").strip().upper()
        caja_input = request.form.get("caja", "").strip().upper()
        coordenadas = request.form["coordenadas"].strip()
        detalles = request.form["detalles"].strip()
        contrata = request.form.get("contrata", "").strip()
        
        # Sede de registro
        target_branch = branch if not es_admin else request.form.get("branch", "ARE").strip().upper()
        
        # Validar XBOX y HUBOX si se proveen
        if xbox and not xbox.startswith("XB"):
            flash("El tipo de caja XBOX debe comenzar con XB.", "danger")
            return redirect(url_for("dashboard"))
        if hubox and not hubox.startswith("HB"):
            flash("El tipo de caja HUBOX debe comenzar con HB.", "danger")
            return redirect(url_for("dashboard"))
            
        # Componer código de caja
        parts = [site]
        if xbox:
            parts.append(xbox)
        if hubox:
            parts.append(hubox)
        if caja_input:
            parts.append(caja_input)
        caja_compuesta = "-".join(parts)
        
        # Validar si ya existe la cuenta (solo si no es un placeholder autogenerado)
        existente = None
        if not cuenta.startswith("SIN_CUENTA_") and not cuenta.startswith("AVERÍA_ODN_"):
            existente = Averia.query.filter_by(cuenta=cuenta, estado="PENDIENTE").order_by(Averia.id.desc()).first()
        if existente:
            if accion == "registrar_y_reparar":
                flash(f"La cuenta {cuenta} ya tiene una avería pendiente activa. Redirigiendo a resolución.", "info")
                return redirect(url_for("dashboard", auto_resolver_id=existente.id, auto_resolver_cuenta=existente.cuenta))
            flash(f"La cuenta {cuenta} ya tiene una avería pendiente activa.", "warning")
            return redirect(url_for("dashboard"))
        
        nueva = Averia(
            branch=target_branch,
            cuenta=cuenta,
            site=site,
            caja=caja_compuesta,
            coordenadas=coordenadas,
            detalles=detalles,
            contrata=contrata or "",
            estado="PENDIENTE",
            dias_pendientes=0.0,
            origen="MANUAL",
            fecha_creacion=obtener_hora_peru()
        )
        db.session.add(nueva)
        db.session.commit()
        flash(f"Avería manual para cuenta {cuenta} creada con éxito.", "success")
        if accion == "registrar_y_reparar":
            return redirect(url_for("dashboard", auto_resolver_id=nueva.id, auto_resolver_cuenta=nueva.cuenta))
    except Exception as e:
        db.session.rollback()
        flash(f"Error al crear avería: {str(e)}", "danger")
        
    return redirect(url_for("dashboard"))


@app.route("/averias/editar/<int:id>", methods=["GET", "POST"])
@login_requerido
def editar_averia(id):
    branch_op = session.get("operador_branch")
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    
    if es_noc:
        return jsonify({"success": False, "message": "No tienes permisos para editar la avería."}), 403
        
    averia = db.session.get(Averia, id)
    if not averia:
        return jsonify({"success": False, "message": "La avería no existe."}), 404
        
    if request.method == "GET":
        # Separar la caja compuesta en sus componentes
        site = averia.site or ""
        xbox = ""
        hubox = ""
        caja_num = ""
        
        if averia.caja:
            parts = averia.caja.split("-")
            if len(parts) > 0:
                site = parts[0]
            for p in parts[1:]:
                if p.upper().startswith("XB"):
                    xbox = p
                elif p.upper().startswith("HB"):
                    hubox = p
                else:
                    caja_num = p
                    
        return jsonify({
            "success": True,
            "id": averia.id,
            "cuenta": averia.cuenta,
            "branch": averia.branch,
            "site": site,
            "xbox": xbox,
            "hubox": hubox,
            "caja_num": caja_num,
            "coordenadas": averia.coordenadas or "",
            "detalles": averia.detalles or "",
            "origen": averia.origen,
            "estado": averia.estado,
            "tipificacion": averia.tipificacion or ""
        })
        
    # POST - Guardar cambios
    try:
        cuenta = request.form.get("cuenta", "").strip()
        if not cuenta:
            flash("La cuenta no puede estar vacía.", "danger")
            return redirect(request.referrer or url_for("dashboard"))
            
        site = request.form["site"].strip().upper()
        xbox = request.form.get("xbox", "").strip().upper()
        hubox = request.form.get("hubox", "").strip().upper()
        caja_input = request.form.get("caja_num", "").strip().upper()
        coordenadas = request.form["coordenadas"].strip()
        detalles = request.form["detalles"].strip()
        contrata = request.form.get("contrata", "").strip()
        
        # Sede (sólo admins pueden cambiarla, u operarios si no hay restricción. Dejémoslo libre si no es NOC)
        target_branch = request.form.get("branch", averia.branch).strip().upper()
        
        # Validar XBOX y HUBOX si se proveen
        if xbox and not xbox.startswith("XB"):
            flash("El tipo de caja XBOX debe comenzar con XB.", "danger")
            return redirect(request.referrer or url_for("dashboard"))
        if hubox and not hubox.startswith("HB"):
            flash("El tipo de caja HUBOX debe comenzar con HB.", "danger")
            return redirect(request.referrer or url_for("dashboard"))
            
        # Componer código de caja
        parts = [site]
        if xbox:
            parts.append(xbox)
        if hubox:
            parts.append(hubox)
        if caja_input:
            parts.append(caja_input)
        caja_compuesta = "-".join(parts)
        
        # Actualizar avería
        averia.cuenta = cuenta
        averia.site = site
        averia.caja = caja_compuesta
        averia.coordenadas = coordenadas
        averia.detalles = detalles
        averia.contrata = contrata or ""
        averia.branch = target_branch
        
        # Si la avería está resuelta, permitir actualizar su tipificación
        if averia.estado == "REPARADO":
            averia.tipificacion = request.form.get("tipificacion", "").strip()
        
        db.session.commit()
        flash("La información de la avería ha sido actualizada con éxito.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error al editar la avería: {str(e)}", "danger")
        
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/stock", methods=["GET", "POST"])
@login_requerido
def registrar_stock():
    branch = session.get("operador_branch")
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    
    if es_admin or es_noc:
        sede_actual = request.args.get("branch", "ARE").strip().upper()
    else:
        sede_actual = branch.strip().upper()
        
    if request.method == "POST":
        if es_noc:
            flash("No tienes permisos para modificar el inventario.", "danger")
            return redirect(url_for("registrar_stock", branch=sede_actual))
        codigos = request.form.getlist("material_codigo")
        nombres = request.form.getlist("material_nombre")
        actuales = request.form.getlist("stock_actual")
        enviados = request.form.getlist("stock_enviado_noc")
        fechas = request.form.getlist("fecha_envio_noc")
        
        for i, nombre in enumerate(nombres):
            cod = codigos[i] if len(codigos) > i else ""
            if not nombre:
                continue
                
            reg = StockBranch.query.filter_by(branch=sede_actual, material_nombre=nombre).first()
            if not reg:
                reg = StockBranch(
                    branch=sede_actual,
                    material_codigo=cod,
                    material_nombre=nombre
                )
                db.session.add(reg)
                
            try:
                submitted_actual = int(actuales[i]) if actuales[i] else 0
            except ValueError:
                submitted_actual = 0
                
            try:
                submitted_enviado = int(enviados[i]) if enviados[i] else 0
            except ValueError:
                submitted_enviado = 0
                
            reg.stock_actual = submitted_actual
            
            # Solo actualizar los detalles del último envío NOC si se ingresó un valor > 0
            if submitted_enviado > 0:
                reg.stock_enviado_noc = submitted_enviado
                fecha_str = fechas[i] if len(fechas) > i else ""
                if fecha_str:
                    try:
                        reg.fecha_envio_noc = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    except ValueError:
                        reg.fecha_envio_noc = datetime.today().date()
                else:
                    reg.fecha_envio_noc = datetime.today().date()

                
        try:
            db.session.commit()
            flash("Stock actualizado correctamente.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error al actualizar stock: {str(e)}", "danger")
            
        return redirect(url_for("registrar_stock", branch=sede_actual))
        
    stock_regs = StockBranch.query.filter_by(branch=sede_actual).all()
    stock_dict = {r.material_nombre: r for r in stock_regs}
    
    materiales_stock = []
    for m in MATERIALES_MASTER:
        reg = stock_dict.get(m["nombre"])
        materiales_stock.append({
            "codigo": m["codigo"],
            "nombre": m["nombre"],
            "seccion": m["seccion"],
            "stock_actual": reg.stock_actual if reg else 0,
            "stock_enviado_noc": reg.stock_enviado_noc if reg else 0,
            "fecha_envio_noc": reg.fecha_envio_noc.strftime("%Y-%m-%d") if reg and reg.fecha_envio_noc else ""
        })
        
    sedes = ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"]
    
    materiales_por_seccion_stock = {}
    for m in materiales_stock:
        sec = m["seccion"]
        if sec not in materiales_por_seccion_stock:
            materiales_por_seccion_stock[sec] = []
        materiales_por_seccion_stock[sec].append(m)
        
    # Calcular consumos de esta sede
    reparadas_sede = Averia.query.filter_by(branch=sede_actual, estado="REPARADO").all()
    consumo_dict = {f"{m['codigo']}|{m['nombre']}": 0 for m in MATERIALES_MASTER}
    
    for av in reparadas_sede:
        mats_dict = av.materiales_dict
        for key, cant in mats_dict.items():
            if cant:
                if key in consumo_dict:
                    consumo_dict[key] += cant
                else:
                    parts = key.split("|")
                    if len(parts) > 0:
                        code = parts[0]
                        found = next((m for m in MATERIALES_MASTER if m["codigo"] == code), None)
                        if found:
                            master_key = f"{found['codigo']}|{found['nombre']}"
                            consumo_dict[master_key] += cant

    consumo_materiales = []
    for m in MATERIALES_MASTER:
        key = f"{m['codigo']}|{m['nombre']}"
        cant = consumo_dict.get(key, 0)
        if cant > 0:
            consumo_materiales.append({
                "codigo": m["codigo"],
                "nombre": m["nombre"],
                "seccion": m["seccion"],
                "cantidad": cant
            })
            
    # Calcular tipificaciones de esta sede
    typifications = [
        "Personas externa cortó",
        "Camión grande rompió fibra",
        "Puerto sucio",
        "Refusión de hilo / cambio de módulo",
        "Cambio de caja",
        "Caja robada",
        "Reemplazo de poste eléctrico",
        "Conector roto",
        "OLT caída"
    ]
    tipificaciones_sede = []
    for typ in typifications:
        count = sum(1 for av in reparadas_sede if av.tipificacion == typ)
        tipificaciones_sede.append({"tipificacion": typ, "cantidad": count})
        
    reparadas_sede_serialized = []
    for av in reparadas_sede:
        reparadas_sede_serialized.append({
            "fecha": av.fecha_resolucion.strftime("%Y-%m-%d") if av.fecha_resolucion else "",
            "cable": av.material_cable_m or 0,
            "conectores": av.material_conectores or 0,
            "rosetas": av.material_rosetas or 0,
            "mangas": av.material_mangas or 0,
            "acopladores": av.material_acopladores or 0,
            "tipificacion": av.tipificacion or "",
            "materiales_dict": av.materiales_dict
        })

    return render_template(
        "stock.html",
        materiales_por_seccion=materiales_por_seccion_stock,
        sede_actual=sede_actual,
        sedes=sedes,
        es_admin_or_noc=(es_admin or es_noc),
        es_noc=es_noc,
        consumo_materiales=consumo_materiales,
        tipificaciones_data=tipificaciones_sede,
        total_reparadas_sede=len(reparadas_sede),
        reparadas_sede_raw=reparadas_sede_serialized
    )


@app.route("/stock/exportar/<branch>", methods=["GET"])
@login_requerido
def exportar_inventario_sede(branch):
    user_branch = session.get("operador_branch")
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    
    branch = branch.strip().upper()
    if not es_admin and not es_noc and user_branch != branch:
        flash("No tienes permisos para descargar el inventario de esta sede.", "danger")
        return redirect(url_for("registrar_stock", branch=user_branch))
        
    stock_regs = StockBranch.query.filter_by(branch=branch).all()
    stock_dict = {r.material_nombre: r for r in stock_regs}
    
    wb = Workbook()
    wb.remove(wb.active)
    
    if es_admin or es_noc:
        sedes_to_export = ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"]
    else:
        sedes_to_export = [branch]
        
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=11)
    
    headers = [
        "Sección",
        "Código",
        "Material",
        "Stock Actual",
        "Último Stock Enviado NOC",
        "Último Fecha Envío NOC"
    ]
    
    for br in sedes_to_export:
        stock_regs = StockBranch.query.filter_by(branch=br).all()
        stock_dict = {r.material_nombre: r for r in stock_regs}
        
        ws = wb.create_sheet(title=br)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        row_idx = 2
        for m in MATERIALES_MASTER:
            reg = stock_dict.get(m["nombre"])
            stock_actual = reg.stock_actual if reg else 0
            enviado_noc = reg.stock_enviado_noc if reg else 0
            fecha_noc = reg.fecha_envio_noc.strftime("%Y-%m-%d") if reg and reg.fecha_envio_noc else ""
            
            row_values = [
                m["seccion"],
                m["codigo"],
                m["nombre"],
                stock_actual,
                enviado_noc,
                fecha_noc
            ]
            
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                if col_idx in [1, 2, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            row_idx += 1
            
        ws.row_dimensions[1].height = 28
        
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"reporte_almacenes_{obtener_hora_peru().strftime('%Y%m%d')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/avance-diario", methods=["GET"])
@login_requerido
def avance_diario():
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    user_branch = session.get("operador_branch")
    
    can_filter_branch = es_admin or es_noc or user_branch == "ALL"
    
    branch_arg = request.args.get("branch")
    if can_filter_branch:
        target_branch = branch_arg if branch_arg else "ALL"
    else:
        target_branch = user_branch
        
    hoy_str = obtener_hora_peru().strftime("%Y-%m-%d")
    date_arg = request.args.get("date", hoy_str).strip()
    
    from sqlalchemy.orm import joinedload
    query = Averia.query.filter_by(estado="REPARADO").options(joinedload(Averia.tecnico))
    if target_branch != "ALL":
        query = query.filter_by(branch=target_branch)
        
    all_resolved = query.order_by(Averia.fecha_resolucion.desc()).all()
    
    matching_averias = []
    for av in all_resolved:
        if av.fecha_resolucion and av.fecha_resolucion.strftime("%Y-%m-%d") == date_arg:
            matching_averias.append(av)
            
    # Calculate stats
    total_cable = 0
    total_conectores = 0
    total_rosetas = 0
    total_mangas = 0
    total_acopladores = 0
    
    for av in matching_averias:
        total_cable += av.material_cable_m or 0
        total_conectores += av.material_conectores or 0
        total_rosetas += av.material_rosetas or 0
        total_mangas += av.material_mangas or 0
        total_acopladores += av.material_acopladores or 0
        
    stats = {
        "count": len(matching_averias),
        "total_cable": total_cable,
        "total_conectores": total_conectores,
        "total_rosetas": total_rosetas,
        "total_mangas": total_mangas,
        "total_acopladores": total_acopladores,
    }
    
    sedes = ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"]
    
    view_arg = request.args.get("view", "data").strip().lower()
    if view_arg not in ["data", "sumario"]:
        view_arg = "data"
        
    # Group by Branch (Sede) for the SUMARIO report
    branch_agg = {}
    for av in matching_averias:
        branch = av.branch
        if branch not in branch_agg:
            branch_agg[branch] = {
                "branch": branch,
                "sites": set(),
                "cajas": set(),
                "clientes": 0
            }
        if av.site:
            branch_agg[branch]["sites"].add(av.site)
        if av.caja:
            branch_agg[branch]["cajas"].add(av.caja)
        branch_agg[branch]["clientes"] += 1
        
    sumario_cajas = []
    for b, data in branch_agg.items():
        sorted_sites = sorted(list(data["sites"]))
        sumario_cajas.append({
            "branch": b,
            "sites_str": ", ".join(sorted_sites) if sorted_sites else "N/A",
            "cajas_count": len(data["cajas"]),
            "clientes_count": data["clientes"]
        })
    sumario_cajas.sort(key=lambda x: x["branch"])
    
    return render_template(
        "avance_diario.html",
        averias=matching_averias,
        sumario_cajas=sumario_cajas,
        stats=stats,
        selected_date=date_arg,
        selected_branch=target_branch,
        can_filter_branch=can_filter_branch,
        sedes=sedes,
        selected_view=view_arg
    )


@app.route("/avance-diario/exportar", methods=["GET"])
@login_requerido
def exportar_avance_diario():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    user_branch = session.get("operador_branch")
    
    can_filter_branch = es_admin or es_noc or user_branch == "ALL"
    
    branch_arg = request.args.get("branch")
    if can_filter_branch:
        target_branch = branch_arg if branch_arg else "ALL"
    else:
        target_branch = user_branch
        
    hoy_str = obtener_hora_peru().strftime("%Y-%m-%d")
    date_arg = request.args.get("date", hoy_str).strip()
    
    from sqlalchemy.orm import joinedload
    query = Averia.query.filter_by(estado="REPARADO").options(joinedload(Averia.tecnico))
    if target_branch != "ALL":
        query = query.filter_by(branch=target_branch)
        
    all_resolved = query.order_by(Averia.fecha_resolucion.desc()).all()
    
    matching_averias = []
    for av in all_resolved:
        if av.fecha_resolucion and av.fecha_resolucion.strftime("%Y-%m-%d") == date_arg:
            matching_averias.append(av)
            
    view_arg = request.args.get("view", "data").strip().lower()
    
    # Calculate stats
    total_cable = 0
    total_conectores = 0
    total_rosetas = 0
    total_mangas = 0
    total_acopladores = 0
    
    for av in matching_averias:
        total_cable += av.material_cable_m or 0
        total_conectores += av.material_conectores or 0
        total_rosetas += av.material_rosetas or 0
        total_mangas += av.material_mangas or 0
        total_acopladores += av.material_acopladores or 0
        
    wb = Workbook()
    
    # Styles
    purple_fill = PatternFill("solid", fgColor="5B21B6") # Deep purple header
    purple_font = Font(color="FFFFFF", bold=True)
    light_purple_fill = PatternFill("solid", fgColor="F5F3FF")
    
    cyan_fill = PatternFill("solid", fgColor="ECFEFF") # Light cyan accent
    cyan_font = Font(color="0891B2", bold=True)
    
    bold_font = Font(bold=True)
    italic_font = Font(italic=True, color="64748B")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    def write_header_and_summary(ws, sheet_title):
        # Title Block
        ws.merge_cells("A1:G1")
        ws["A1"] = sheet_title
        ws["A1"].font = Font(size=14, bold=True, color="5B21B6")
        ws["A1"].alignment = left_align
        ws.row_dimensions[1].height = 30
        
        ws["A2"] = f"Fecha de reporte: {date_arg}"
        ws["A2"].font = bold_font
        ws["A3"] = f"Sede consultada: {target_branch if target_branch != 'ALL' else 'TODAS LAS SEDES'}"
        ws["A3"].font = bold_font
        ws["A4"] = f"Generado el: {obtener_hora_peru().strftime('%Y-%m-%d %H:%M')}"
        ws["A4"].font = italic_font
        
        # Summary Block
        ws["A6"] = "RESUMEN DE CONSUMO"
        ws["A6"].font = cyan_font
        ws["A7"] = "Métrica"
        ws["B7"] = "Valor"
        ws["A7"].font = bold_font
        ws["B7"].font = bold_font
        ws["A7"].fill = light_purple_fill
        ws["B7"].fill = light_purple_fill
        
        summary_metrics = [
            ("Averías Resueltas", len(matching_averias), "ud"),
            ("Cable Drop", total_cable, "m"),
            ("FAC", total_conectores, "ud"),
            ("Waterproof", total_rosetas, "ud"),
            ("Mufas", total_mangas, "ud"),
            ("Preconectorizado", total_acopladores, "ud")
        ]
        
        for idx, (label, val, unit) in enumerate(summary_metrics, 8):
            ws[f"A{idx}"] = label
            ws[f"B{idx}"] = f"{val} {unit}"
            ws[f"A{idx}"].alignment = left_align
            ws[f"B{idx}"].alignment = right_align

    def write_details_table(ws, start_row):
        ws[f"A{start_row - 1}"] = "DETALLE DE AVERÍAS"
        ws[f"A{start_row - 1}"].font = cyan_font
        
        headers = [
            "Sede", "Cuenta / Cliente", "Tipo", "Caja", "SITE", 
            "WO", "Detalles / Comentarios", "Técnico Resuelve", "DNI Técnico",
            "Cable Drop (m)", "FAC (ud)", "Waterproof (ud)", "Mufas (ud)", "Preconectorizado (ud)",
            "Hora Resolución"
        ]
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.fill = purple_fill
            cell.font = purple_font
            cell.alignment = center_align
            
        ws.row_dimensions[start_row].height = 28
        
        row_idx = start_row + 1
        for av in matching_averias:
            tipo = "Agrupado" if av.material_comentarios and "Agrupado en la avería principal" in av.material_comentarios else "Principal"
            tech_name = av.tecnico.nombre if av.tecnico else "No asignado"
            tech_dni = av.tecnico.dni if av.tecnico else ""
            hora_resol = av.fecha_resolucion.strftime('%I:%M %p') if av.fecha_resolucion else ""
            
            row_values = [
                av.branch,
                av.cuenta or "",
                tipo,
                av.caja or "",
                av.site or "",
                av.codigo_wo or "",
                av.detalles or "",
                tech_name,
                tech_dni,
                av.material_cable_m or 0,
                av.material_conectores or 0,
                av.material_rosetas or 0,
                av.material_mangas or 0,
                av.material_acopladores or 0,
                hora_resol
            ]
            
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(size=10)
                if col_idx in [1, 3, 4, 5, 6, 9, 15]:
                    cell.alignment = center_align
                elif col_idx in [10, 11, 12, 13, 14]:
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align
            row_idx += 1

    if view_arg == "sumario":
        # Sheet 1: Sumario
        ws_sumario = wb.active
        ws_sumario.title = "Sumario"
        write_header_and_summary(ws_sumario, "AVANCE DIARIO - SUMARIO DE CAJAS REPARADAS")
        
        # Group by Branch (Sede)
        branch_agg = {}
        for av in matching_averias:
            branch = av.branch
            if branch not in branch_agg:
                branch_agg[branch] = {
                    "branch": branch,
                    "sites": set(),
                    "cajas": set(),
                    "clientes": 0
                }
            if av.site:
                branch_agg[branch]["sites"].add(av.site)
            if av.caja:
                branch_agg[branch]["cajas"].add(av.caja)
            branch_agg[branch]["clientes"] += 1
            
        sumario_cajas = []
        for b, data in branch_agg.items():
            sorted_sites = sorted(list(data["sites"]))
            sumario_cajas.append({
                "branch": b,
                "sites_str": ", ".join(sorted_sites) if sorted_sites else "N/A",
                "cajas_count": len(data["cajas"]),
                "clientes_count": data["clientes"]
            })
        sumario_cajas.sort(key=lambda x: x["branch"])
        
        start_row = 16
        ws_sumario[f"A{start_row - 1}"] = "SUMARIO DE CAJAS Y CLIENTES POR SEDE"
        ws_sumario[f"A{start_row - 1}"].font = cyan_font
        
        headers = [
            "Sede", "Sites Reparados", "Cantidad de Cajas", "Cantidad de Clientes Afectados"
        ]
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws_sumario.cell(row=start_row, column=col_idx, value=h)
            cell.fill = purple_fill
            cell.font = purple_font
            cell.alignment = center_align
            
        ws_sumario.row_dimensions[start_row].height = 28
        
        row_idx = start_row + 1
        for item in sumario_cajas:
            row_values = [
                item["branch"],
                item["sites_str"],
                item["cajas_count"],
                item["clientes_count"]
            ]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws_sumario.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(size=10)
                if col_idx in [1, 3, 4]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            row_idx += 1
            
        # Sheet 2: Cuentas Reparadas (Detalle)
        ws_detalle = wb.create_sheet(title="Cuentas Reparadas")
        write_header_and_summary(ws_detalle, "AVANCE DIARIO - REPORTES DE AVERÍAS RESUELTAS")
        write_details_table(ws_detalle, start_row=16)
        
    else:
        # Sheet 1: Detalle Cuentas
        ws_detalle = wb.active
        ws_detalle.title = "Cuentas Reparadas"
        write_header_and_summary(ws_detalle, "AVANCE DIARIO - REPORTES DE AVERÍAS RESUELTAS")
        write_details_table(ws_detalle, start_row=16)
            
    # Auto fit column width for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"avance_diario_{view_arg}_{target_branch}_{date_arg}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/averias", methods=["GET"])
@login_requerido
def listar_averias():
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    branch = session.get("operador_branch")
    
    from sqlalchemy.orm import joinedload
    query = Averia.query.options(joinedload(Averia.tecnico))
    if not es_admin and not es_noc and branch != "ALL":
        query = query.filter_by(branch=branch)
        
    averias = query.order_by(Averia.id.desc()).all()
    stats = obtener_estadisticas(branch=branch, es_admin=(es_admin or es_noc))
    return render_template("averias.html", averias=averias, stats=stats)


@app.route("/averias/exportar", methods=["GET"])
@login_requerido
def exportar_averias():
    import collections
    es_admin = session.get("operador_rol") == "admin"
    es_noc = session.get("operador_rol") == "noc"
    user_branch = session.get("operador_branch")
    
    branch_arg = request.args.get("branch")
    month_arg = request.args.get("month")
    site_arg = request.args.get("site")
    status_arg = request.args.get("status")
    origin_arg = request.args.get("origin")
    q_arg = request.args.get("q")
    
    if not es_admin and not es_noc and user_branch != "ALL":
        target_branch = user_branch
    else:
        target_branch = branch_arg if branch_arg else "ALL"
        
    from sqlalchemy.orm import joinedload
    query = Averia.query.options(joinedload(Averia.tecnico))
    if target_branch != "ALL":
        query = query.filter_by(branch=target_branch)
        
    averias = query.order_by(Averia.id.desc()).all()
    
    # Filter by month python-side (consistent with frontend data-month)
    if month_arg:
        filtered_averias = []
        for av in averias:
            av_date = av.fecha_resolucion if av.fecha_resolucion else av.fecha_creacion
            if av_date and av_date.strftime("%Y-%m") == month_arg:
                filtered_averias.append(av)
        averias = filtered_averias

    # Filter by site
    if site_arg:
        averias = [av for av in averias if av.site == site_arg]
        
    # Filter by origin
    if origin_arg:
        averias = [av for av in averias if av.origen == origin_arg]
        
    # Filter by status
    if status_arg:
        if status_arg == "PENDIENTE":
            averias = [av for av in averias if av.estado == "PENDIENTE"]
        elif status_arg == "REPARADO":
            averias = [av for av in averias if av.estado == "REPARADO"]
        elif status_arg == "REPARADO_PRINCIPAL":
            averias = [av for av in averias if av.estado == "REPARADO" and not (av.material_comentarios and "Agrupado en la avería principal" in av.material_comentarios)]
        elif status_arg == "REPARADO_AGRUPADO":
            averias = [av for av in averias if av.estado == "REPARADO" and av.material_comentarios and "Agrupado en la avería principal" in av.material_comentarios]
            
    # Filter by search query (q)
    if q_arg:
        q_clean = q_arg.lower().strip()
        filtered_q = []
        for av in averias:
            tecnico_name = av.tecnico.nombre.lower() if av.tecnico else ""
            search_fields = [
                av.cuenta or "",
                av.caja or "",
                av.site or "",
                av.codigo_wo or "",
                av.detalles or "",
                tecnico_name
            ]
            search_str = " ".join(search_fields).lower()
            if q_clean in search_str:
                filtered_q.append(av)
        averias = filtered_q
        
    reparadas = [av for av in averias if av.estado == "REPARADO" and av.fecha_resolucion]
    
    # Re-structure materials into a lookup dict for extremely fast O(1) lookup
    mats_lookup_global = {}
    for av in averias:
        mats_dict = av.materiales_dict
        by_key = {}
        by_code = collections.defaultdict(int)
        for k, c in mats_dict.items():
            if c is None:
                continue
            try:
                qty = int(c)
            except ValueError:
                qty = 0
            by_key[k] = qty
            if "|" in k:
                parts = k.split("|", 1)
                code = parts[0]
                if code and code != "Sin Código":
                    by_code[code] += qty
        mats_lookup_global[av.id] = {
            'by_key': by_key,
            'by_code': by_code
        }
    
    # Group reparadas by month (YYYY-MM)
    reparadas_por_mes = collections.defaultdict(list)
    for av in reparadas:
        mes_str = av.fecha_resolucion.strftime("%Y-%m")
        reparadas_por_mes[mes_str].append(av)
        
    # Group reparadas by month and branch
    reparadas_por_mes_y_branch = collections.defaultdict(list)
    for av in reparadas:
        mes_str = av.fecha_resolucion.strftime("%Y-%m")
        br = av.branch or "Sin Sede"
        reparadas_por_mes_y_branch[(mes_str, br)].append(av)
        
    # Build dictionary of stock by branch and material name
    stock_by_branch = collections.defaultdict(dict)
    all_stock = StockBranch.query.all()
    for r in all_stock:
        stock_by_branch[r.branch][r.material_nombre] = {
            "stock_actual": r.stock_actual or 0,
            "stock_enviado_noc": r.stock_enviado_noc or 0,
            "fecha_envio_noc": r.fecha_envio_noc.strftime("%Y-%m-%d") if r.fecha_envio_noc else ""
        }
        
    workbook = Workbook()
    
    # Styling definitions
    purple_fill = PatternFill("solid", fgColor="5B21B6") # Deep purple header
    purple_font = Font(color="FFFFFF", bold=True)
    
    cyan_fill = PatternFill("solid", fgColor="ECFEFF") # Light cyan accent
    cyan_font = Font(color="0891B2", bold=True)
    
    green_fill = PatternFill("solid", fgColor="D1FAE5") # Light green total
    green_font = Font(color="065F46", bold=True)
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    def get_material_quantity(lookup, material_obj):
        key = f"{material_obj['codigo']}|{material_obj['nombre']}"
        cant = lookup['by_key'].get(key)
        if cant is not None:
            return cant
        code = material_obj['codigo']
        if code and code != "Sin Código":
            return lookup['by_code'].get(code, 0)
        return 0

    sin_reporte_raw = request.args.get("sin_reporte_raw") == "true"
    months_keys = sorted(list(reparadas_por_mes.keys()), reverse=True)

    def populate_summary_tables(ws, title_prefix, branch_filter):
        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = title_prefix
        ws["A1"].font = Font(size=14, bold=True, color="5B21B6")
        ws["A1"].alignment = left_align
        ws.row_dimensions[1].height = 30
        
        ws["A2"] = f"Generado el: {obtener_hora_peru().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(italic=True, color="64748B")
        
        curr_row = 4
        
        # Filter reparadas for this branch
        if branch_filter:
            reparadas_scope = [av for av in reparadas if av.branch == branch_filter]
        else:
            reparadas_scope = reparadas
            
        # Pre-aggregate monthly material consumption and tipificaciones in a single loop
        # to avoid executing millions of nested loop query steps.
        material_consumption = collections.defaultdict(int)
        tipificacion_counts = collections.defaultdict(int)
        
        for av in reparadas_scope:
            mes_str = av.fecha_resolucion.strftime("%Y-%m")
            lookup = mats_lookup_global[av.id]
            for m in MATERIALES_MASTER:
                cant = get_material_quantity(lookup, m)
                if cant > 0:
                    material_consumption[(m["nombre"], mes_str)] += cant
            if av.tipificacion:
                tipificacion_counts[(av.tipificacion, mes_str)] += 1
            
        # Table 1: Material Consumption
        ws.cell(row=curr_row, column=1, value="RESUMEN DE CONSUMO DE MATERIALES").font = Font(size=11, bold=True, color="0891B2")
        curr_row += 1
        
        headers_t1 = ["Material", "Código"] + months_keys + ["Total Consumido"]
        for col_idx, h in enumerate(headers_t1, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=h)
            cell.fill = purple_fill
            cell.font = purple_font
            cell.alignment = center_align
        
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1
        
        for m in MATERIALES_MASTER:
            has_consumption = False
            row_values = []
            for mes in months_keys:
                cant = material_consumption[(m["nombre"], mes)]
                row_values.append(cant)
                if cant > 0:
                    has_consumption = True
                    
            if not has_consumption:
                continue
                
            ws.cell(row=curr_row, column=1, value=m["nombre"]) # default left-align is fine
            ws.cell(row=curr_row, column=2, value=m["codigo"]).alignment = center_align
            
            for idx, cant in enumerate(row_values):
                # Excel naturally right-aligns numeric values. Writing value directly is fast.
                ws.cell(row=curr_row, column=3 + idx, value=cant)
                
            total_col_idx = 3 + len(months_keys)
            total_cell = ws.cell(row=curr_row, column=total_col_idx)
            total_cell.font = bold_font
            total_cell.fill = green_fill
            if months_keys:
                start_letter = get_column_letter(3)
                end_letter = get_column_letter(2 + len(months_keys))
                total_cell.value = f"=SUM({start_letter}{curr_row}:{end_letter}{curr_row})"
            else:
                total_cell.value = 0
            
            curr_row += 1
            
        curr_row += 2
        
        # Table 2: Typifications
        ws.cell(row=curr_row, column=1, value="RESUMEN DE TIPIFICACIONES DE AVERÍAS").font = Font(size=11, bold=True, color="0891B2")
        curr_row += 1
        
        typifications = [
            "Personas externa cortó",
            "Camión grande rompió fibra",
            "Puerto sucio",
            "Refusión de hilo / cambio de módulo",
            "Cambio de caja",
            "Caja robada",
            "Reemplazo de poste eléctrico",
            "Conector roto",
            "OLT caída"
        ]
        
        headers_t2 = ["Tipificación"] + months_keys + ["Total Incidencias"]
        for col_idx, h in enumerate(headers_t2, start=1):
            cell = ws.cell(row=curr_row, column=col_idx, value=h)
            cell.fill = purple_fill
            cell.font = purple_font
            cell.alignment = center_align
            
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1
        
        for typ in typifications:
            ws.cell(row=curr_row, column=1, value=typ) # default left-align is fine
            
            row_counts = []
            for mes in months_keys:
                count = tipificacion_counts[(typ, mes)]
                row_counts.append(count)
                
            for idx, count in enumerate(row_counts):
                # Excel naturally right-aligns numeric values.
                ws.cell(row=curr_row, column=2 + idx, value=count)
                
            total_col_idx = 2 + len(months_keys)
            total_cell = ws.cell(row=curr_row, column=total_col_idx)
            total_cell.font = bold_font
            total_cell.fill = green_fill
            if months_keys:
                start_letter = get_column_letter(2)
                end_letter = get_column_letter(1 + len(months_keys))
                total_cell.value = f"=SUM({start_letter}{curr_row}:{end_letter}{curr_row})"
            else:
                total_cell.value = 0
            
            curr_row += 1

    # 1. SUMMARY SHEET
    summary_sheet = workbook.active
    summary_sheet.title = "SUMMARY"
    populate_summary_tables(summary_sheet, "DASHBOARD GENERAL DE CONSUMO E INCIDENCIAS", None)
    
    # 2. RAW DATA SHEET ("Reporte Averias ODN") if not excluded
    if not sin_reporte_raw:
        raw_sheet = workbook.create_sheet(title="Reporte Averias ODN")
        
        headers_raw = [
            "ID",
            "Sede (Branch)",
            "Código de WO",
            "Cuenta",
            "Detalles",
            "Días Pendientes",
            "Estado del Ticket",
            "Contrata",
            "Site",
            "Caja",
            "Coordenadas",
            "Origen",
            "Fecha Creación",
            "Fecha Resolución",
            "Técnico que Resolvió",
            "Tipificación",
            "Cuentas Agrupadas"
        ]
        
        for m in MATERIALES_MASTER:
            headers_raw.append(f"[{m['codigo']}] {m['nombre']}")
            
        headers_raw.append("Comentarios Solución")
        raw_sheet.append(headers_raw)
        
        curr_raw_row = 2
        for av in averias:
            tecnico_nombre = av.tecnico.nombre if av.tecnico else ""
            row_data = [
                av.id,
                av.branch,
                av.codigo_wo or "",
                av.cuenta or "",
                av.detalles or "",
                av.dias_pendientes or 0.0,
                av.estado,
                av.contrata or "",
                av.site or "",
                av.caja or "",
                av.coordenadas or "",
                av.origen,
                av.fecha_creacion.strftime("%Y-%m-%d %H:%M") if av.fecha_creacion else "",
                av.fecha_resolucion.strftime("%Y-%m-%d %H:%M") if av.fecha_resolucion else "Pendiente",
                tecnico_nombre,
                av.tipificacion or "",
                av.cuentas_asociadas or ""
            ]
            
            lookup = mats_lookup_global[av.id]
            for m in MATERIALES_MASTER:
                cant = get_material_quantity(lookup, m)
                row_data.append(cant if cant > 0 else "")
                
            row_data.append(av.material_comentarios or "")
            
            # Write only non-empty cells to avoid openpyxl Cell overhead
            for col_idx, val in enumerate(row_data, 1):
                if val != "" and val is not None:
                    raw_sheet.cell(row=curr_raw_row, column=col_idx, value=val)
            curr_raw_row += 1
            
        raw_header_fill = PatternFill("solid", fgColor="1D4ED8")
        raw_header_font = Font(color="FFFFFF", bold=True)
        
        for cell in raw_sheet[1]:
            cell.fill = raw_header_fill
            cell.font = raw_header_font
            cell.alignment = center_align
            
        raw_sheet.freeze_panes = "A2"
        raw_sheet.auto_filter.ref = raw_sheet.dimensions
        
        for row in raw_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 3. INDIVIDUAL BRANCH SHEETS
    if target_branch == "ALL":
        active_branches = set(av.branch for av in reparadas if av.branch)
        branches_to_generate = [br for br in ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"] if br in active_branches]
    else:
        branches_to_generate = []
        
    for br in branches_to_generate:
        br_sheet = workbook.create_sheet(title=br)
        populate_summary_tables(br_sheet, f"DASHBOARD DE CONSUMO E INCIDENCIAS - Sede: {br}", br)

    # Global column auto-fit (optimized utilizing values_only iter_rows hash mapping, capped at 100 rows to prevent timeouts on large datasets)
    for ws in workbook.worksheets:
        col_widths = {}
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            if row_count > 100:
                break
            for col_idx, val in enumerate(row, 1):
                val_str = str(val or "")
                if val_str.startswith("="):
                    val_str = "0.00"
                length = len(val_str)
                if col_idx not in col_widths or length > col_widths[col_idx]:
                    col_widths[col_idx] = length
        for col_idx, max_len in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

    import gc
    gc.collect()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Construct descriptive filename
    filename_parts = ["reporte_averias"]
    if target_branch != "ALL":
        filename_parts.append(target_branch.lower())
    else:
        filename_parts.append("global")
    if month_arg:
        filename_parts.append(month_arg)
    download_name = "_".join(filename_parts) + ".xlsx"
    
    return send_file(
        output,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/noc")
@login_requerido
def noc_dashboard():
    if session.get("operador_rol") not in ["noc", "admin"]:
        flash("Acceso denegado: Se requieren permisos de NOC o Administrador.", "danger")
        return redirect(url_for("dashboard"))
        
    reparadas = Averia.query.filter_by(estado="REPARADO").order_by(Averia.fecha_resolucion.desc()).all()
    
    total_cable = 0
    total_conectores = 0
    total_rosetas = 0
    total_mangas = 0
    total_acopladores = 0
    
    MESES_ES = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    
    por_mes = {}
    por_branch = {}
    
    for av in reparadas:
        fecha = av.fecha_resolucion
        if fecha:
            mes_key = fecha.strftime("%Y-%m")
            mes_label = f"{MESES_ES.get(fecha.month, 'Desconocido')} {fecha.year}"
        else:
            mes_key = "Sin Fecha"
            mes_label = "Sin Fecha"
            
        br = av.branch or "Sin Sede"
        
        cable = av.material_cable_m or 0
        conectores = av.material_conectores or 0
        rosetas = av.material_rosetas or 0
        mangas = av.material_mangas or 0
        acopladores = av.material_acopladores or 0
        
        total_cable += cable
        total_conectores += conectores
        total_rosetas += rosetas
        total_mangas += mangas
        total_acopladores += acopladores
        
        # Agrupación por Mes
        if mes_key not in por_mes:
            por_mes[mes_key] = {
                "key": mes_key,
                "label": mes_label,
                "cable": 0,
                "conectores": 0,
                "rosetas": 0,
                "mangas": 0,
                "acopladores": 0,
                "total_casos": 0
            }
        por_mes[mes_key]["cable"] += cable
        por_mes[mes_key]["conectores"] += conectores
        por_mes[mes_key]["rosetas"] += rosetas
        por_mes[mes_key]["mangas"] += mangas
        por_mes[mes_key]["acopladores"] += acopladores
        por_mes[mes_key]["total_casos"] += 1
        
        # Agrupación por Branch (Sede)
        if br not in por_branch:
            por_branch[br] = {
                "branch": br,
                "cable": 0,
                "conectores": 0,
                "rosetas": 0,
                "mangas": 0,
                "acopladores": 0,
                "total_casos": 0
            }
        por_branch[br]["cable"] += cable
        por_branch[br]["conectores"] += conectores
        por_branch[br]["rosetas"] += rosetas
        por_branch[br]["mangas"] += mangas
        por_branch[br]["acopladores"] += acopladores
        por_branch[br]["total_casos"] += 1

    # Ordenar agrupaciones
    por_mes_lista = sorted(por_mes.values(), key=lambda x: x["key"], reverse=True)
    por_branch_lista = sorted(por_branch.values(), key=lambda x: x["branch"])
    
    stats_noc = {
        "total_reparadas": len(reparadas),
        "total_cable": total_cable,
        "total_conectores": total_conectores,
        "total_rosetas": total_rosetas,
        "total_mangas": total_mangas,
        "total_acopladores": total_acopladores
    }
    
    typifications = [
        "Personas externa cortó",
        "Camión grande rompió fibra",
        "Puerto sucio",
        "Refusión de hilo / cambio de módulo",
        "Cambio de caja",
        "Caja robada",
        "Reemplazo de poste eléctrico",
        "Conector roto",
        "OLT caída"
    ]
    tipificaciones_data = []
    for typ in typifications:
        count = sum(1 for av in reparadas if av.tipificacion == typ)
        tipificaciones_data.append({"tipificacion": typ, "cantidad": count})
        
    stock_regs = StockBranch.query.order_by(StockBranch.branch, StockBranch.material_codigo).all()
    reparadas_raw = []
    for av in reparadas:
        reparadas_raw.append({
            "branch": av.branch or "Sin Sede",
            "fecha": av.fecha_resolucion.strftime("%Y-%m-%d") if av.fecha_resolucion else "",
            "cable": av.material_cable_m or 0,
            "conectores": av.material_conectores or 0,
            "rosetas": av.material_rosetas or 0,
            "mangas": av.material_mangas or 0,
            "acopladores": av.material_acopladores or 0,
            "tipificacion": av.tipificacion or "",
            "materiales_dict": av.materiales_dict
        })
    sedes = ["LI1", "LI2", "LI3", "LI4", "LI7", "ARE", "PIU", "SAN", "CAJ", "LAL", "HUN", "CUS", "JUN"]

    return render_template(
        "noc_dashboard.html",
        reparadas=reparadas,
        por_mes=por_mes_lista,
        por_branch=por_branch_lista,
        stats=stats_noc,
        tipificaciones_data=tipificaciones_data,
        stock_regs=stock_regs,
        reparadas_raw=reparadas_raw,
        sedes=sedes
    )


def migrar_comentarios_agrupacion():
    try:
        averias_agrupadas = Averia.query.filter(
            Averia.material_comentarios.like("Agrupado en la avería principal%")
        ).all()
        
        migrated_count = 0
        import re
        for av in averias_agrupadas:
            comm = av.material_comentarios
            m_id = re.search(r"ID\s+(\d+)", comm)
            m_cuenta = re.search(r"Cuenta\s+([^)]+)", comm)
            
            main_cuenta = None
            if m_id:
                main_id = int(m_id.group(1))
                main_av = db.session.get(Averia, main_id)
                if main_av:
                    main_cuenta = main_av.cuenta
            elif m_cuenta:
                main_cuenta = m_cuenta.group(1).strip()
                
            if main_cuenta:
                new_comm = f"Agrupado en la avería principal ({main_cuenta})"
                if comm != new_comm:
                    av.material_comentarios = new_comm
                    migrated_count += 1
                    
        if migrated_count > 0:
            db.session.commit()
            print(f"Migración de comentarios de agrupación completada. Se actualizaron {migrated_count} registros.")
    except Exception as e:
        db.session.rollback()
        print("Error en migración de comentarios de agrupación:", e)


def migrar_materiales_mapeados():
    try:
        reparadas = Averia.query.filter_by(estado="REPARADO").all()
        migrated_count = 0
        import json
        for av in reparadas:
            if av.materiales_json and av.materiales_json != "{}":
                try:
                    mats = json.loads(av.materiales_json)
                except Exception:
                    continue
                
                cable_m = 0
                conectores = 0
                rosetas = 0
                mangas = 0
                acopladores = 0
                
                for key, cant in mats.items():
                    parts = key.split("|")
                    codigo = parts[0]
                    nombre = parts[1] if len(parts) > 1 else ""
                    
                    m_cod, m_nom, m_sec = obtener_material_mapeado(codigo, nombre)
                    
                    if m_nom == "Cable Drop":
                        cable_m += cant
                    elif m_nom == "FAC":
                        conectores += cant
                    elif m_nom == "Waterproof":
                        rosetas += cant
                    elif m_nom == "Mufas":
                        mangas += cant
                    elif m_nom == "Preconectorizado":
                        acopladores += cant
                
                # Verificar si los valores actuales difieren
                if (av.material_cable_m != cable_m or 
                    av.material_conectores != conectores or 
                    av.material_rosetas != rosetas or 
                    av.material_mangas != mangas or 
                    av.material_acopladores != acopladores):
                    
                    av.material_cable_m = cable_m
                    av.material_conectores = conectores
                    av.material_rosetas = rosetas
                    av.material_mangas = mangas
                    av.material_acopladores = acopladores
                    migrated_count += 1
                    
        if migrated_count > 0:
            db.session.commit()
            print(f"Migración de mapeo de materiales completada. Se actualizaron {migrated_count} registros.")
    except Exception as e:
        db.session.rollback()
        print("Error en migración de mapeo de materiales:", e)


def migrar_recuperar_materiales_perdidos():
    try:
        # Buscar todas las averías REPARADO que no tienen materiales
        reparadas_sin_mats = Averia.query.filter(
            Averia.estado == "REPARADO",
            (Averia.material_cable_m.is_(None) | (Averia.material_cable_m == 0)),
            (Averia.material_conectores.is_(None) | (Averia.material_conectores == 0)),
            (Averia.material_rosetas.is_(None) | (Averia.material_rosetas == 0)),
            (Averia.material_mangas.is_(None) | (Averia.material_mangas == 0)),
            (Averia.material_acopladores.is_(None) | (Averia.material_acopladores == 0)),
            (Averia.materiales_json.is_(None) | (Averia.materiales_json == "{}") | (Averia.materiales_json == ""))
        ).all()
        
        migrated_count = 0
        for av in reparadas_sin_mats:
            # Intentar buscar y copiar materiales previos de la misma cuenta
            if buscar_y_copiar_materiales_previos(av.cuenta, av):
                migrated_count += 1
                
        if migrated_count > 0:
            db.session.commit()
            print(f"Migración: Se recuperaron los materiales de {migrated_count} averías a partir de registros históricos.")
    except Exception as e:
        db.session.rollback()
        print("Error en migración de recuperación de materiales perdidos:", e)


def migrar_nombres_grillete_clevis():
    try:
        import json
        mapping_grillete = {
            "Grillete tipo D": "Clevis tipo D",
            "Grillete tipo trébol": "Clevis tipo trébol",
        }
        
        # 1. Migrar StockBranch
        stock_updated = 0
        stock_deleted = 0
        for old_name, new_name in mapping_grillete.items():
            records = StockBranch.query.filter_by(material_nombre=old_name).all()
            for rec in records:
                exist_rec = StockBranch.query.filter_by(branch=rec.branch, material_nombre=new_name).first()
                if exist_rec:
                    exist_rec.stock_actual += rec.stock_actual
                    exist_rec.stock_enviado_noc += rec.stock_enviado_noc
                    db.session.delete(rec)
                    stock_deleted += 1
                else:
                    rec.material_nombre = new_name
                    stock_updated += 1
                    
        # 2. Migrar Averia.materiales_json
        averia_updated = 0
        averias = Averia.query.all()
        for av in averias:
            changed = False
            if av.materiales_json:
                try:
                    mats = json.loads(av.materiales_json)
                    new_mats = {}
                    for key, val in mats.items():
                        if "|" in key:
                            parts = key.split("|", 1)
                            code = parts[0]
                            name = parts[1]
                            if name in mapping_grillete:
                                new_key = f"{code}|{mapping_grillete[name]}"
                                new_mats[new_key] = val
                                changed = True
                            else:
                                new_mats[key] = val
                        else:
                            if key in mapping_grillete:
                                new_mats[mapping_grillete[key]] = val
                                changed = True
                            else:
                                new_mats[key] = val
                    if changed:
                        av.materiales_json = json.dumps(new_mats)
                except Exception:
                    pass
            
            if av.material_comentarios:
                old_com = av.material_comentarios
                new_com = old_com.replace("Grillete", "Clevis").replace("grillete", "clevis").replace("GRILLETE", "CLEVIS")
                if new_com != old_com:
                    av.material_comentarios = new_com
                    changed = True
                    
            if changed:
                averia_updated += 1
                
        if stock_updated > 0 or stock_deleted > 0 or averia_updated > 0:
            db.session.commit()
            print(f"Migración Clevis: {stock_updated} stocks actualizados, {stock_deleted} fusionados, {averia_updated} averías migradas.")
    except Exception as e:
        db.session.rollback()
        print("Error en migración de nombres Grillete/Clevis:", e)


def limpiar_contrata_tgi():
    try:
        updated = Averia.query.filter(
            (Averia.contrata == "TGI") | (Averia.contrata == "Sin contrata")
        ).update({"contrata": ""}, synchronize_session=False)
        if updated > 0:
            db.session.commit()
            print(f"Migración: Se limpiaron {updated} averías que tenían contrata 'TGI' o 'Sin contrata'.")
    except Exception as e:
        db.session.rollback()
        print("Error en migración limpiar_contrata_tgi:", e)


def migrar_nombres_vano_span():
    try:
        import json
        mapping_names = {
            "Cable de fibra óptica ADSS 48 fibras, vano 100m": "Fibra óptica ADSS 48 fibras, SPAN 100m",
            "Cable de fibra óptica ADSS 24 fibras, vano 200m": "Fibra óptica ADSS 24 fibras, SPAN 200m",
            "Cable de fibra óptica ADSS 24 fibras, vano 100m": "Fibra óptica ADSS 24 fibras, SPAN 100m",
            "Cable de fibra óptica ASU 12Fo, vano 100m": "Fibra óptica ASU 12Fo, SPAN 100m",
            "Cable de fibra óptica ASU 4Fo, vano 100m": "Fibra óptica ASU 4Fo, SPAN 100m",
            "Cable de fibra óptica ASU 4Fo, vano 100m - Flexible": "Fibra óptica ASU 4Fo, SPAN 100m - Flexible",
            "Cable de fibra óptica ASU 1Fo, vano 100m": "Fibra óptica ASU 1Fo, SPAN 100m",
            "Cable de fibra óptica 1 hilo - Drop": "Fibra óptica 1 hilo - Drop",
            "Grapa de tensión para vano de 200m": "Grapa de tensión para SPAN de 200m",
            "Grapa de tensión para vano de 100m": "Grapa de tensión para SPAN de 100m",
            "Grapa de suspensión para vano de 100m": "Grapa de suspensión para SPAN de 100m",
            "Etiqueta para cable de fibra óptica": "Etiqueta para fibra óptica",
        }
        
        # 1. Migrar StockBranch
        stock_updated = 0
        stock_deleted = 0
        for old_name, new_name in mapping_names.items():
            records = StockBranch.query.filter_by(material_nombre=old_name).all()
            for rec in records:
                # Verificar si ya existe un registro con el nuevo nombre para esta misma sede
                exist_rec = StockBranch.query.filter_by(branch=rec.branch, material_nombre=new_name).first()
                if exist_rec:
                    exist_rec.stock_actual += rec.stock_actual
                    exist_rec.stock_enviado_noc += rec.stock_enviado_noc
                    db.session.delete(rec)
                    stock_deleted += 1
                else:
                    rec.material_nombre = new_name
                    stock_updated += 1
                    
        # 2. Migrar Averia.materiales_json
        averia_updated = 0
        averias = Averia.query.all()
        for av in averias:
            if av.materiales_json:
                try:
                    mats = json.loads(av.materiales_json)
                    new_mats = {}
                    changed = False
                    for key, val in mats.items():
                        if "|" in key:
                            parts = key.split("|", 1)
                            code = parts[0]
                            name = parts[1]
                            if name in mapping_names:
                                new_key = f"{code}|{mapping_names[name]}"
                                new_mats[new_key] = val
                                changed = True
                            else:
                                new_mats[key] = val
                        else:
                            if key in mapping_names:
                                new_mats[mapping_names[key]] = val
                                changed = True
                            else:
                                new_mats[key] = val
                    if changed:
                        av.materiales_json = json.dumps(new_mats)
                        averia_updated += 1
                except Exception:
                    pass
                    
        if stock_updated > 0 or stock_deleted > 0 or averia_updated > 0:
            db.session.commit()
            print(f"Migración SPAN/Fibra: {stock_updated} stocks actualizados, {stock_deleted} fusionados, {averia_updated} averías migradas.")
    except Exception as e:
        db.session.rollback()
        print("Error en migración de nombres SPAN/Fibra:", e)


with app.app_context():
    try:
        db.create_all()
        asegurar_esquema()
        crear_operador_defecto()
        limpiar_contrata_tgi()
        migrar_comentarios_agrupacion()
        migrar_materiales_mapeados()
        migrar_recuperar_materiales_perdidos()
        migrar_nombres_vano_span()
        migrar_nombres_grillete_clevis()
        # Verify if static/sites.json exists, otherwise fetch it
        import os
        sites_path = os.path.join(app.root_path, "static", "sites.json")
        if not os.path.exists(sites_path):
            print("SITES cache not found, fetching...")
            sincronizar_sites()
            
        # Verify if boxes table has records, otherwise fetch it
        try:
            if Box.query.first() is None:
                print("Boxes table is empty, fetching from Google Sheets...")
                sincronizar_boxes()
        except Exception as box_err:
            print("Error checking or populating boxes on startup:", box_err)
            
        print("Base de datos y esquemas inicializados correctamente.")
    except Exception as e:
        print("Error inicializando base de datos en arranque:", e)


if __name__ == "__main__":
    app.run(port=5000)
